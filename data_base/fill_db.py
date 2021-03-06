import openpyxl
import psycopg2
import time
import math
import threading
from data_base.postgresql import Postgresql
from werkzeug.security import generate_password_hash
from glob import glob
from configurations.development import Config
from vpr_analysis.vpr_2022 import VPR22
from vpr_analysis.vpr_analysis import VPR
from vpr_analysis.vpr_2020 import VPR20


class FillDb(Postgresql):
    def __init__(self, connection):
        super().__init__(connection)

    def create_index_on_result_for_task(self):
        self._cur.execute("CREATE INDEX ON result_for_task (id_oo_parallels_subjects, id_oo_parallels );")

    def create_index_on_result_for_task_distributio_of_tasks_by_positions_of_codifiers(self):
        self._cur.execute(
            "CREATE INDEX ON result_for_task_distributio_of_tasks_by_positions_of_codifiers (id_subjects, parallel);")

    def create_index_on_oo_parallels_subjects(self):
        self._cur.execute(
            "CREATE INDEX ON oo_parallels_subjects (id_subjects, id_oo_parallels);")

    def get_id_kt(self, kt_key, id_subjects, parallel, year):
        try:
            if not kt_key:
                kt_key = 'NULL'
            self._cur.execute(f"""
            SELECT id_kt FROM kt 
            WHERE kt_key = '{str(kt_key).strip()}' 
            and id_subjects = {id_subjects} 
            and parallel = {parallel}
            and year = '{year}';""")
            res = self._cur.fetchone()
            if res:
                return res[0]
            print("kt", kt_key, self.get_subject_name(id_subjects), parallel)
            return None
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_ks(self, ks_key, id_subjects, parallel, year):
        try:
            if not ks_key:
                ks_key = 'NULL'
            self._cur.execute(f"""
                SELECT id_ks FROM ks 
                WHERE ks_key = '{str(ks_key).strip()}' 
                and id_subjects = {id_subjects} 
                and parallel = {parallel}
                and year = '{year}';""")
            res = self._cur.fetchone()
            if res:
                return res[0]
            print("ks", ks_key, self.get_subject_name(id_subjects), parallel)
            return None
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_organizational_and_legal_form(self, type_of_organizational_and_legal_form):
        try:
            type_ = type_of_organizational_and_legal_form[3:].strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_organizational_and_legal_form FROM organizational_and_legal_form 
            WHERE type_of_organizational_and_legal_form = '{type_}';""")
            res = self._cur.fetchone()
            if not res:
                print(type_)
                print("id_organizational_and_legal_form не найден")
                return []
            return res[0]
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_name_of_the_settlement(self, id_district, id_oo_location_type, name):
        try:
            self._cur.execute(f"""
            SELECT id_name_of_the_settlement FROM name_of_the_settlement 
            WHERE id_district = {id_district} 
            AND id_oo_location_type = {id_oo_location_type} 
            AND name = '{name}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_name_of_the_settlement не найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_districts(self, district_name):
        try:
            district_name = district_name.replace(' ', '_').replace('-', '_')
            self._cur.execute(f"""
            SELECT id_district FROM district 
            WHERE district_name = '{district_name}';""")
            res, = self._cur.fetchone()
            if res:
                return res
        except psycopg2.Error as e:
            print("Ошибка получения id_districts из БД " + str(e))
        return []

    def get_id_population_of_the_settlement(self, interval):
        try:
            interval = interval[3:].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_population_of_the_settlement FROM population_of_the_settlement 
            WHERE interval = '{interval}';""")

            res = self._cur.fetchone()
            if not res:
                print(interval)
                print("id_population_of_the_settlement не был найден")
                return []
            return res[0]
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_internet_speed(self, interval):
        try:
            interval = interval.strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_internet_speed FROM internet_speed 
            WHERE interval = '{interval}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_internet_speed не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_count_of_parents_attending_events(self, description):
        try:
            if description is None:
                description = "4.Нет_данных"
            description = description[2:].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(',
                                                                                                               '(')
            self._cur.execute(f"""
            SELECT id_count_of_parents_attending_events FROM count_of_parents_attending_events 
            WHERE description = '{description}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_count_of_parents_attending_events не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_count_of_parents_ready_to_help(self, description):
        try:
            if description is None:
                description = "4.Нет_данных"
            description = description[2:].strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_count_of_parents_ready_to_help FROM count_of_parents_ready_to_help 
            WHERE description = '{description}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_count_of_parents_ready_to_help не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_regular_transport_link(self, description):
        try:
            if description is None:
                description = "5. Нет_данных"
            description = description[3:].strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_regular_transport_link FROM regular_transport_link 
            WHERE description = '{description}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_regular_transport_link не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_frequency_of_regular_transport_link(self, description):
        try:
            if description is None:
                description = "5. Нет_данных"
            description = description[3:].strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_frequency_of_regular_transport_link FROM frequency_of_regular_transport_link 
            WHERE description = '{description}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_frequency_of_regular_transport_link не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_possibility_to_get_to_the_oo_by_public_transport(self, description):
        try:
            if description is None:
                description = "5. Нет_данных"
            description = description[3:].strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(',
                                                            '(')
            self._cur.execute(f"""
            SELECT id_possibility_to_get_to_the_oo_by_public_transport FROM possibility_to_get_to_the_oo_by_public_transport 
            WHERE description = '{description}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_possibility_to_get_to_the_oo_by_public_transport не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_the_involvement_of_students_in_additional_education(self, interval) -> list:
        try:
            if interval is None:
                interval = "5. Нет даннных"
            interval = interval[3:].strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_the_involvement_of_students_in_additional_education FROM the_involvement_of_students_in_additional_education 
            WHERE interval = '{interval}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_internet_speed не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_subjects(self, subject_name):
        try:
            subject_name = subject_name.replace(" ", "_")
            self._cur.execute(f"""
            SELECT id_subjects FROM subjects 
            WHERE subject_name = '{subject_name}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_subjects не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_classes(self, id_oo_parallels, liter):
        try:
            self._cur.execute(f"""
            SELECT id_classes FROM classes 
            WHERE id_oo_parallels = {id_oo_parallels} 
            AND liter = '{liter}';""")
            res, = self._cur.fetchone()
            if not res:
                print("id_classes не был найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_textbooks(self, id_subjects, name):
        try:
            self._cur.execute(
                f"""
                SELECT id_textbooks FROM textbooks 
                WHERE id_subjects = {id_subjects} 
                AND name = '{name}';""")
            res = self._cur.fetchone()
            if not res:
                print("id_textbooks не был найден")
                print(id_subjects, name)
                self._cur.execute(
                    f"""
                    SELECT id_textbooks FROM textbooks 
                    WHERE id_subjects = {id_subjects} 
                    AND name = '{'Нет данных'}';""")
                res = self._cur.fetchone()
            res, = res
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_oo_login(self, id_oo):
        try:
            self._cur.execute(f"""
            SELECT oo_login FROM oo 
            WHERE id_oo = {id_oo};""")
            res, = self._cur.fetchone()
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_students(self, id_oo_parallels, id_classes, student_number):
        try:
            self._cur.execute(f"""
            SELECT id_students FROM students 
            WHERE id_oo_parallels = {id_oo_parallels} 
            AND id_classes = {id_classes} 
            AND student_number = '{student_number}';""")
            res, = self._cur.fetchone()
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def get_id_distributio_of_tasks_by_positions_of_codifiers(self, id_subjects, parallel, task_number, year):
        try:
            self._cur.execute(f"""
            SELECT id_distributio_of_tasks_by_positions_of_codifiers 
            FROM distributio_of_tasks_by_positions_of_codifiers 
            WHERE id_subjects = {id_subjects} 
            AND parallel = {parallel} 
            AND task_number = {task_number}
            AND year = '{year}';""")
            res = self._cur.fetchone()
            if res:
                return res[0]
            return None
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def create_roles(self):
        try:
            self._cur.execute(f"""
            INSERT INTO roles (role) 
            VALUES ('admin'), ('ministry'), ('municipality'), ('school')""")
            print("Roles created")
        except psycopg2.Error as e:
            print("Ошибка: " + str(e))

    def create_users(self):
        try:
            users_wb = openpyxl.reader.excel.load_workbook(
                filename="../excel/users.xlsx", data_only=True)
            users_sheet = users_wb.active
            for row in range(2, users_sheet.max_row + 1):
                login = users_sheet["B" + str(row)].value
                name = users_sheet["C" + str(row)].value
                hash_psw = generate_password_hash(users_sheet["F" + str(row)].value)
                tm = math.floor(time.time())
                id_role = int(users_sheet["H" + str(row)].value)
                self._cur.execute(f"""
                INSERT INTO users (login, name, password, id_role, time) 
                VALUES ('{login}', '{name}', '{hash_psw}', {id_role}, {tm})""")
            print("Users created")
        except psycopg2.Error as e:
            print("Ошибка: " + str(e))

    def dropAllTables(self):
        try:
            self._cur.execute("DROP SCHEMA public CASCADE;")
            print("All Tables Drop")
        except psycopg2.Error as e:
            print("Ошибка: " + str(e))

    def createTables(self):
        try:
            self._cur.execute("CREATE SCHEMA IF NOT EXISTS public AUTHORIZATION vpr_user;")
            self._cur.execute(open("../sql/Create_Tables.sql", "r").read())
            print("Tables Created")
        except psycopg2.Error as e:
            print("Ошибка: " + str(e))

    def fill_district(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/logins.xlsx", data_only=True)
            data_sheet = data.active
            district_list = []
            for row in range(2, data_sheet.max_row + 1):
                district_value = data_sheet["C" + str(row)].value
                if district_value not in district_list:
                    district_list.append(district_value)
            for district_name in sorted(district_list):
                self._cur.execute(f"""
                INSERT INTO district (district_name) 
                VALUES ('{district_name.replace(' ', '_').replace('-', '_')}')""")
            print("Таблица district заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def find_district_id_by_district_name(self, district_name):
        try:
            district_name = district_name.replace(' ', '_').replace('-', '_')
            self._cur.execute(f"""
            SELECT id_district FROM district 
            WHERE district_name = '{district_name}';""")
            res, = self._cur.fetchone()
            if not res:
                print("Район не найден")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def fill_oo_location_type(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            oo_location_type_list = []
            for row in range(2, data_sheet.max_row + 1):
                oo_location_type = data_sheet["N" + str(row)].value
                add_list = [oo_location_type[0], oo_location_type[2:]]
                if add_list not in oo_location_type_list:
                    oo_location_type_list.append(add_list)
            for location_type in sorted(oo_location_type_list, key=lambda x: x[0]):
                location_type = location_type[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace(
                    '_(', '(')
                self._cur.execute(f"""
                INSERT INTO oo_location_type (location_type) 
                VALUES ('{location_type}')""")
            print("Таблица oo_location_type заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def find_oo_location_type_id_by_location_type(self, location_type):
        try:
            location_type = location_type[2:]
            location_type = location_type.strip().replace(' ', '_') \
                .replace('-', '_').replace(',', '').replace('_(', '(')
            self._cur.execute(f"""
            SELECT id_oo_location_type FROM oo_location_type 
            WHERE location_type = '{location_type}';""")
            res, = self._cur.fetchone()
            if not res:
                print("Не найден тип расположения ОО")
                return []
            return res
        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def fill_name_of_the_settlement(self, year):
        try:
            if year == 2021:
                data = openpyxl.reader.excel.load_workbook(
                    filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
                data_sheet = data.active

                login_data = openpyxl.reader.excel.load_workbook(
                    filename="../excel/logins.xlsx", data_only=True)
                login_data_sheet = login_data.active
                data_list = []
                for data_row in range(2, data_sheet.max_row + 1):
                    login_value = data_sheet["A" + str(data_row)].value
                    for login_row in range(2, login_data_sheet.max_row + 1):
                        if login_data_sheet["A" + str(login_row)].value == login_value:
                            district_id = self.find_district_id_by_district_name(
                                login_data_sheet["C" + str(login_row)].value)
                            oo_location_type_id = self.find_oo_location_type_id_by_location_type(
                                data_sheet["N" + str(data_row)].value)
                            name = data_sheet["O" + str(data_row)].value
                            add_list = [district_id, oo_location_type_id, name]
                            if add_list not in data_list:
                                data_list.append(add_list)
                for lst in data_list:
                    id_district, id_oo_location_type, name = lst
                    self._cur.execute(f"""
                    INSERT INTO name_of_the_settlement (id_district, id_oo_location_type, name) 
                    VALUES ({id_district}, {id_oo_location_type}, '{name}');""")
                data.close()
                login_data.close()
            if year == 2022:
                # sch526312
                id_district = self.get_id_districts("Павловский район")
                id_oo_location_type = 2
                name = "г.Павлово"
                self._cur.execute(f"""
                insert into name_of_the_settlement (id_district, id_oo_location_type, name) 
                values ({id_district}, {id_oo_location_type}, '{name}')
                """)
                # sch526312
                id_district = 3
                id_oo_location_type = 2
                name = "г. Камень-на-Оби"
                self._cur.execute(f"""
                                insert into name_of_the_settlement (id_district, id_oo_location_type, name) 
                                values ({id_district}, {id_oo_location_type}, '{name}')
                                """)
            print("Таблица name_of_the_settlement заполненна ")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_organizational_and_legal_form(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            organizational_and_legal_form_list = []
            for row in range(2, data_sheet.max_row + 1):
                organizational_and_legal_form = data_sheet["I" + str(row)].value
                add_list = [organizational_and_legal_form[0], organizational_and_legal_form[2:]]
                if add_list not in organizational_and_legal_form_list:
                    organizational_and_legal_form_list.append(add_list)
            for type_of_organizational_and_legal_form in sorted(organizational_and_legal_form_list, key=lambda x: x[0]):
                type_of_organizational_and_legal_form = type_of_organizational_and_legal_form[1].strip().replace(' ',
                                                                                                                 '_').replace(
                    '-', '_').replace(',', '').replace('_(', '(')
                self._cur.execute(f"""
                INSERT INTO organizational_and_legal_form (type_of_organizational_and_legal_form) 
                VALUES ('{type_of_organizational_and_legal_form}');""")
            print("Таблица organizational_and_legal_form заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_oo_logins(self, year):
        try:
            if year == 2021:
                data = openpyxl.reader.excel.load_workbook(
                    filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
                data_sheet = data.active
                oo_logins_list = []
                for row in range(2, data_sheet.max_row + 1):
                    oo_logins = data_sheet["A" + str(row)].value
                    if oo_logins not in oo_logins_list:
                        oo_logins_list.append(oo_logins)
                for login in oo_logins_list:
                    self._cur.execute(f"""
                    INSERT INTO oo_logins (oo_login) 
                    VALUES ('{login.strip()}');""")
                print("Таблица oo_logins заполненна")
                data.close()
            if year == 2022:
                data = openpyxl.reader.excel.load_workbook(
                    filename="../excel/vpr_results/2022/Сбор контекстных данных об ОО (2022).xlsx", data_only=True)
                sheet_1 = data.active
                oo_logins_list = []
                for row in range(2, sheet_1.max_row + 1):
                    oo_logins: str = sheet_1["A" + str(row)].value
                    oo_logins = oo_logins.replace("edu", "sch")
                    if oo_logins not in oo_logins_list:
                        oo_logins_list.append(oo_logins)
                for login in oo_logins_list:
                    try:
                        self._cur.execute(f"select oo_login from oo_logins where oo_login = '{login}';")
                        res = self._cur.fetchone()
                        if not res:
                            self._cur.execute(f"insert into oo_logins (oo_login) values ('{login}')")
                            print(f"add {login}")
                            # sch526312
                            # ksh220032
                    except psycopg2.Error as e:
                        print("Ошибка при заполнении БД " + str(e))

        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_population_of_the_settlement(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            population_of_the_settlement_list = []
            for row in range(2, data_sheet.max_row + 1):
                population_of_the_settlement = data_sheet["P" + str(row)].value
                add_list = [population_of_the_settlement[0], population_of_the_settlement[2:]]
                if add_list not in population_of_the_settlement_list:
                    population_of_the_settlement_list.append(add_list)
            for interval in sorted(population_of_the_settlement_list, key=lambda x: x[0]):
                self._cur.execute(
                    f"INSERT INTO population_of_the_settlement (interval) VALUES ('"
                    f"{interval[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица population_of_the_settlement заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_internet_speed(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            internet_speed_list = []
            for row in range(2, data_sheet.max_row + 1):
                internet_speed = data_sheet["BR" + str(row)].value
                if internet_speed not in internet_speed_list:
                    internet_speed_list.append(internet_speed)
            internet_speed_list[1], internet_speed_list[3] = internet_speed_list[3], internet_speed_list[1]
            internet_speed_list[2], internet_speed_list[3] = internet_speed_list[3], internet_speed_list[2]

            for interval in internet_speed_list:
                self._cur.execute(
                    f"INSERT INTO internet_speed (interval) VALUES ('"
                    f"{interval.strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица internet_speed заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_the_involvement_of_students_in_additional_education(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            the_involvement_of_students_in_additional_education_list = []
            for row in range(2, data_sheet.max_row + 1):
                the_involvement_of_students_in_additional_education = data_sheet["BS" + str(row)].value
                if the_involvement_of_students_in_additional_education not in the_involvement_of_students_in_additional_education_list:
                    the_involvement_of_students_in_additional_education_list.append(
                        the_involvement_of_students_in_additional_education)

            for interval in the_involvement_of_students_in_additional_education_list:
                if interval is None:
                    interval = "5. Нет даннных"
                self._cur.execute(
                    f"INSERT INTO the_involvement_of_students_in_additional_education (interval) VALUES ('"
                    f"{interval[3:].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица the_involvement_of_students_in_additional_education заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_count_of_parents_attending_events(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            count_of_parents_attending_events_list = []
            for row in range(2, data_sheet.max_row + 1):
                count_of_parents_attending_events = data_sheet["CD" + str(row)].value
                if count_of_parents_attending_events is None:
                    count_of_parents_attending_events = "4. Нет_данных"
                add_list = [count_of_parents_attending_events[0], count_of_parents_attending_events[2:]]
                if add_list not in count_of_parents_attending_events_list:
                    count_of_parents_attending_events_list.append(add_list)
            for description in sorted(count_of_parents_attending_events_list, key=lambda x: x[0]):
                self._cur.execute(
                    f"INSERT INTO count_of_parents_attending_events (description) VALUES ('"
                    f"{description[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица count_of_parents_attending_events заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_count_of_parents_ready_to_help(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            count_of_parents_ready_to_help_list = []
            for row in range(2, data_sheet.max_row + 1):
                count_of_parents_ready_to_help = data_sheet["CE" + str(row)].value
                if count_of_parents_ready_to_help is None:
                    count_of_parents_ready_to_help = "4. Нет_данных"
                add_list = [count_of_parents_ready_to_help[0], count_of_parents_ready_to_help[2:]]
                if add_list not in count_of_parents_ready_to_help_list:
                    count_of_parents_ready_to_help_list.append(add_list)
            for description in sorted(count_of_parents_ready_to_help_list, key=lambda x: x[0]):
                self._cur.execute(
                    f"INSERT INTO count_of_parents_ready_to_help (description) VALUES ('"
                    f"{description[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица count_of_parents_ready_to_help заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_regular_transport_link(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            regular_transport_link_list = []
            for row in range(2, data_sheet.max_row + 1):
                regular_transport_link = data_sheet["Q" + str(row)].value
                if regular_transport_link is None:
                    regular_transport_link = "5. Нет_данных"
                add_list = [regular_transport_link[0], regular_transport_link[2:]]
                if add_list not in regular_transport_link_list:
                    regular_transport_link_list.append(add_list)
            for description in sorted(regular_transport_link_list, key=lambda x: x[0]):
                self._cur.execute(
                    f"INSERT INTO regular_transport_link (description) VALUES ('"
                    f"{description[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица regular_transport_link заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_frequency_of_regular_transport_link(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            frequency_of_regular_transport_link_list = []
            for row in range(2, data_sheet.max_row + 1):
                frequency_of_regular_transport_link = data_sheet["R" + str(row)].value
                if frequency_of_regular_transport_link is None:
                    frequency_of_regular_transport_link = "5. Нет_данных"
                add_list = [frequency_of_regular_transport_link[0], frequency_of_regular_transport_link[2:]]
                if add_list not in frequency_of_regular_transport_link_list:
                    frequency_of_regular_transport_link_list.append(add_list)
            for description in sorted(frequency_of_regular_transport_link_list, key=lambda x: x[0]):
                self._cur.execute(
                    f"INSERT INTO frequency_of_regular_transport_link (description) VALUES ('"
                    f"{description[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица frequency_of_regular_transport_link заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_possibility_to_get_to_the_oo_by_public_transport(self):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            possibility_to_get_to_the_oo_by_public_transport_list = []
            for row in range(2, data_sheet.max_row + 1):
                possibility_to_get_to_the_oo_by_public_transport = data_sheet["S" + str(row)].value
                if possibility_to_get_to_the_oo_by_public_transport is None:
                    possibility_to_get_to_the_oo_by_public_transport = "5. Нет_данных"
                add_list = [possibility_to_get_to_the_oo_by_public_transport[0],
                            possibility_to_get_to_the_oo_by_public_transport[2:]]
                if add_list not in possibility_to_get_to_the_oo_by_public_transport_list:
                    possibility_to_get_to_the_oo_by_public_transport_list.append(add_list)
            for description in sorted(possibility_to_get_to_the_oo_by_public_transport_list, key=lambda x: x[0]):
                self._cur.execute(
                    f"INSERT INTO possibility_to_get_to_the_oo_by_public_transport (description) VALUES ('"
                    f"{description[1].strip().replace(' ', '_').replace('-', '_').replace(',', '').replace('_(', '(')}')")
            print("Таблица possibility_to_get_to_the_oo_by_public_transport заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_organisation_status(self):
        statuses = ["1) Самостоятельное юридическое лицо (или головная ОО при наличии филиалов)",
                    "2) Филиал юридического лица, головной ОО"]
        for status in statuses:
            self._cur.execute(f"insert into organisation_status (status) values ('{status}');")
    def get_id_organisation_status(self, status):
        self._cur.execute(f"""
        select id_organisation_status from organisation_status 
        where status = '{status}';""")
        res = self._cur.fetchone()
        if res:
            return res[0]
        return None

    def fill_oo(self, year):
        try:
            if year in (2020, 2021):
                data = openpyxl.reader.excel.load_workbook(
                    filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
                data_sheet = data.active

                login_data = openpyxl.reader.excel.load_workbook(
                    filename="../excel/logins.xlsx", data_only=True)
                login_data_sheet = login_data.active
                login_region = {}
                for row in range(2, login_data_sheet.max_row + 1):
                    login = login_data_sheet["A" + str(row)].value
                    region = login_data_sheet["C" + str(row)].value

                    if login_region.get(region) is None:
                        login_region[region] = [login]
                    else:
                        login_region[region].append(login)
                for row in range(2, data_sheet.max_row + 1):
                    oo_login = data_sheet["A" + str(row)].value
                    for region_d, login_d in login_region.items():
                        if oo_login in login_d:
                            id_name_of_the_settlement = self.get_id_name_of_the_settlement(
                                self.get_id_districts(region_d),
                                self.find_oo_location_type_id_by_location_type(data_sheet["N" + str(row)].value),
                                data_sheet["O" + str(row)].value)
                    id_organizational_and_legal_form = self.get_id_organizational_and_legal_form(
                        data_sheet["I" + str(row)].value)
                    id_population_of_the_settlement = self.get_id_population_of_the_settlement(
                        data_sheet["P" + str(row)].value)
                    id_internet_speed = self.get_id_internet_speed(data_sheet["BR" + str(row)].value)
                    id_the_involvement_of_students_in_additional_education = self.get_id_the_involvement_of_students_in_additional_education(
                        data_sheet["BS" + str(row)].value)
                    id_count_of_parents_attending_events = self.get_id_count_of_parents_attending_events(
                        data_sheet["CD" + str(row)].value)
                    id_count_of_parents_ready_to_help = self.get_id_count_of_parents_ready_to_help(
                        data_sheet["CE" + str(row)].value)
                    id_regular_transport_link = self.get_id_regular_transport_link(data_sheet["Q" + str(row)].value)
                    id_frequency_of_regular_transport_link = self.get_id_frequency_of_regular_transport_link(
                        data_sheet["R" + str(row)].value)
                    id_possibility_to_get_to_the_oo_by_public_transport = self.get_id_possibility_to_get_to_the_oo_by_public_transport(
                        data_sheet["S" + str(row)].value)
                    oo_name = data_sheet["C" + str(row)].value
                    oo_full_name = data_sheet["D" + str(row)].value
                    oo_address = data_sheet["E" + str(row)].value
                    full_name_of_the_director = data_sheet["F" + str(row)].value
                    email_oo = data_sheet["H" + str(row)].value
                    phone_number = str(data_sheet["G" + str(row)].value).replace("(", "").replace(")", "").replace("-",
                                                                                                                   "")
                    self._cur.execute(
                        f"INSERT INTO oo (oo_login,"
                        f" year,"
                        f"id_name_of_the_settlement,"
                        f"id_organizational_and_legal_form,"
                        f"id_population_of_the_settlement,"
                        f"id_internet_speed,"
                        f"id_the_involvement_of_students_in_additional_education,"
                        f"id_count_of_parents_attending_events,"
                        f"id_count_of_parents_ready_to_help,"
                        f"id_regular_transport_link,"
                        f"id_frequency_of_regular_transport_link,"
                        f"id_possibility_to_get_to_the_oo_by_public_transport,"
                        f"oo_name,"
                        f"oo_full_name,"
                        f"oo_address,"
                        f"full_name_of_the_director,"
                        f"email_oo,"
                        f"phone_number) "
                        f"VALUES ('{oo_login}',"
                        f"'{year}',"
                        f"{id_name_of_the_settlement},"
                        f"{id_organizational_and_legal_form},"
                        f"{id_population_of_the_settlement},"
                        f"{id_internet_speed},"
                        f"{id_the_involvement_of_students_in_additional_education},"
                        f"{id_count_of_parents_attending_events},"
                        f"{id_count_of_parents_ready_to_help},"
                        f" {id_regular_transport_link},"
                        f" {id_frequency_of_regular_transport_link},"
                        f" {id_possibility_to_get_to_the_oo_by_public_transport},"
                        f" '{oo_name}',"
                        f" '{oo_full_name}',"
                        f" '{oo_address}',"
                        f" '{full_name_of_the_director}',"
                        f" '{email_oo}',"
                        f" '{phone_number}')")
                data.close()
                login_data.close()
            if year == 2022:
                # data = openpyxl.reader.excel.load_workbook(
                #     filename="../excel/vpr_results/2022/Сбор контекстных данных об ОО (2022).xlsx", data_only=True)
                # data_sheet = data.active
                #
                # login_data = openpyxl.reader.excel.load_workbook(
                #     filename="../excel/logins.xlsx", data_only=True)
                # login_data_sheet = login_data.active
                # login_region = {}
                # for row in range(2, login_data_sheet.max_row + 1):
                #     login = login_data_sheet["A" + str(row)].value
                #     region = login_data_sheet["C" + str(row)].value
                #
                #     if login_region.get(region) is None:
                #         login_region[region] = [login]
                #     else:
                #         login_region[region].append(login)
                # for row in range(2, data_sheet.max_row + 1):
                #     oo_login: str = data_sheet["A" + str(row)].value
                #     oo_login = oo_login.replace("edu", "sch")
                #     for region_d, login_d in login_region.items():
                #         if oo_login in login_d:
                #             self._cur.execute(f"""
                #             select id_name_of_the_settlement from oo
                #             where oo_login = '{oo_login}'
                #             and year = '{2021}';""")
                #             id_name_of_the_settlement = self._cur.fetchone()
                #             if not id_name_of_the_settlement:
                #                 if oo_login == "sch526312":
                #                     id_name_of_the_settlement = 459
                #                 elif oo_login == "ksh220032":
                #                     id_name_of_the_settlement = 460
                #                 else:
                #                     print(oo_login)
                #             else:
                #                 id_name_of_the_settlement = id_name_of_the_settlement[0]
                #     id_organizational_and_legal_form = self.get_id_organizational_and_legal_form(
                #         data_sheet["R" + str(row)].value)
                #     id_population_of_the_settlement = self.get_id_population_of_the_settlement(
                #         data_sheet["V" + str(row)].value)
                #
                #     id_possibility_to_get_to_the_oo_by_public_transport = self.get_id_possibility_to_get_to_the_oo_by_public_transport(
                #         data_sheet["W" + str(row)].value)
                #     oo_name = data_sheet["I" + str(row)].value
                #     oo_full_name = data_sheet["H" + str(row)].value
                #     oo_address = data_sheet["T" + str(row)].value
                #     full_name_of_the_director = data_sheet["J" + str(row)].value
                #     email_oo = data_sheet["L" + str(row)].value
                #     phone_number = str(data_sheet["K" + str(row)].value).replace("(", "").replace(")", "").replace("-","")
                #     inn = data_sheet["M" + str(row)].value
                #     key_oge = data_sheet["N" + str(row)].value
                #     key_ege = data_sheet["O" + str(row)].value
                #     place_index = data_sheet["S" + str(row)].value
                #     id_organisation_status = self.get_id_organisation_status(data_sheet["P" + str(row)].value)
                #     if not id_organisation_status:
                #         print(row, data_sheet["P" + str(row)].value)
                #         continue
                #     try:
                #         self._cur.execute(f"""
                #             INSERT INTO oo (oo_login, year, id_name_of_the_settlement, id_organizational_and_legal_form,
                #             id_population_of_the_settlement, id_possibility_to_get_to_the_oo_by_public_transport,
                #             oo_name, oo_full_name, oo_address, full_name_of_the_director, email_oo, phone_number,
                #             inn, key_oge, key_ege, id_organisation_status, place_index)
                #             VALUES ('{oo_login}', '{year}', {id_name_of_the_settlement}, {id_organizational_and_legal_form},
                #             {id_population_of_the_settlement}, {id_possibility_to_get_to_the_oo_by_public_transport},
                #              '{oo_name}', '{oo_full_name}', '{oo_address}', '{full_name_of_the_director}', '{email_oo}',
                #              '{phone_number}', '{inn}', '{key_oge}', '{key_ege}', {id_organisation_status}, '{place_index}')""")
                #     except psycopg2.Error as e:
                #         print("Ошибка при заполнении БД " + str(e))
                self._cur.execute("""select * from oo where oo_login = 'sch226047' and year= '2021';""")
                res = self._cur.fetchone()
                print(len(res))
                id_oo, oo_login,year,id_name_of_the_settlement, id_organizational_and_legal_form, id_population_of_the_settlement , \
                id_internet_speed, id_the_involvement_of_students_in_additional_education, id_count_of_parents_attending_events, id_count_of_parents_ready_to_help,\
                id_regular_transport_link, id_frequency_of_regular_transport_link, id_possibility_to_get_to_the_oo_by_public_transport, oo_name, \
                oo_full_name, oo_address, full_name_of_the_director, email_oo, phone_number, inn, key_oge, key_ege, id_organisation_status, place_index = res

                self._cur.execute(f"""
                                            INSERT INTO oo (oo_login, year, id_name_of_the_settlement, id_organizational_and_legal_form,
                                            id_population_of_the_settlement, id_possibility_to_get_to_the_oo_by_public_transport,
                                            oo_name, oo_full_name, oo_address, full_name_of_the_director, email_oo, phone_number
                                            )
                                            VALUES ('{oo_login}', '{2022}', {id_name_of_the_settlement}, {id_organizational_and_legal_form},
                                            {id_population_of_the_settlement}, {id_possibility_to_get_to_the_oo_by_public_transport},
                                             '{oo_name}', '{oo_full_name}', '{oo_address}', '{full_name_of_the_director}', '{email_oo}',
                                             '{phone_number}')""")
            print("Таблица ОО заполненна")

        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_duration_of_refresher_courses(self):
        try:
            for values in ["Длительностью_не_более_36_часов",
                           "Длительностью_более_36_часов_и_не_более_72_часов",
                           "Длительностью_более_72_часов"]:
                self._cur.execute(f"INSERT INTO duration_of_refresher_courses (duration) VALUES ('{values}')")
            print("Таблица duration_of_refresher_courses заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_completed_advanced_training_courses_for_teachers(self, year):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            for row in range(2, data_sheet.max_row + 1):
                count_of_teachers_36h = data_sheet["BE" + str(row)].value if data_sheet["BE" + str(row)].value else 0
                count_of_teachers_36h_72h = data_sheet["BF" + str(row)].value if data_sheet[
                    "BF" + str(row)].value else 0
                count_of_teachers_72h = data_sheet["BG" + str(row)].value if data_sheet["BG" + str(row)].value else 0
                oo_login = self.get_id_oo(data_sheet["A" + str(row)].value, year)

                for id_duration, count_of_teachers in zip([1, 2, 3, 4],
                                                          [count_of_teachers_36h, count_of_teachers_36h_72h,
                                                           count_of_teachers_72h]):
                    self._cur.execute(f"INSERT INTO completed_advanced_training_courses_for_teachers ("
                                      f" id_oo,"
                                      f" id_duration_of_refresher_courses,"
                                      f" count_of_teachers)"
                                      f" VALUES ({oo_login}, {id_duration}, {count_of_teachers})")
            print("Таблица completed_advanced_training_courses_for_teachers заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_description_of_work_with_teachers_taking_advanced_training_courses(self):
        try:
            for values in ["Их_сопровождает_методслужба_муниципалитета",
                           "Они_делятся_опытом_в_рамках_профессионального_общения_с_коллегами_в_школе",
                           "Они_самостоятельно_отрабатывают_полученные_навыки_на_уроках",
                           "никак"]:
                self._cur.execute(
                    f"INSERT INTO description_of_work_with_teachers_taking_advanced_training_courses (description) VALUES ('{values}')")
            print("Таблица description_of_work_with_teachers_taking_advanced_training_courses заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_work_with_teachers_taking_advanced_training_courses(self, year):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            for row in range(2, data_sheet.max_row + 1):
                description_1 = data_sheet["BH" + str(row)].value if data_sheet["BH" + str(row)].value else "Нет_данных"
                description_2 = data_sheet["BI" + str(row)].value if data_sheet["BI" + str(row)].value else "Нет_данных"
                description_3 = data_sheet["BJ" + str(row)].value if data_sheet["BJ" + str(row)].value else "Нет_данных"
                description_4 = data_sheet["BK" + str(row)].value if data_sheet["BK" + str(row)].value else "Нет_данных"
                oo_login = self.get_id_oo(data_sheet["A" + str(row)].value, year)
                for id_description, description in zip([1, 2, 3, 4],
                                                       [description_1, description_2, description_3, description_4]):
                    self._cur.execute(f"INSERT INTO completed_advanced_training_courses_for_teachers ("
                                      f" id_oo,"
                                      f" id_duration_of_refresher_courses,"
                                      f" value)"
                                      f" VALUES ({oo_login}, {id_description}, '{description}')")
            print("Таблица work_with_teachers_taking_advanced_training_courses заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_description_of_career_guidance(self):
        try:
            for values in ["Рассказ о профессиях во время классных часов",
                           "Беседы с представителями различных профессий",
                           "Лекции сотрудников службы занятости",
                           "Психологические тестирования построение профессиограмм",
                           "Экскурсии в организации на производства",
                           "Участие в профориентационных проектах",
                           "Практика на предприятиях"]:
                self._cur.execute(
                    f"INSERT INTO description_of_career_guidance (description) VALUES ('{values.replace(' ', '_')}')")
            print("Таблица description_of_career_guidance заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_oo_description_of_career_guidance(self, year):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            for row in range(2, data_sheet.max_row + 1):
                description_1 = data_sheet["BU" + str(row)].value if data_sheet["BU" + str(row)].value else "Нет_данных"
                description_2 = data_sheet["BV" + str(row)].value if data_sheet["BV" + str(row)].value else "Нет_данных"
                description_3 = data_sheet["BW" + str(row)].value if data_sheet["BW" + str(row)].value else "Нет_данных"
                description_4 = data_sheet["BX" + str(row)].value if data_sheet["BX" + str(row)].value else "Нет_данных"
                description_5 = data_sheet["BY" + str(row)].value if data_sheet["BY" + str(row)].value else "Нет_данных"
                description_6 = data_sheet["BZ" + str(row)].value if data_sheet["BZ" + str(row)].value else "Нет_данных"
                description_7 = data_sheet["CA" + str(row)].value if data_sheet["CA" + str(row)].value else "Нет_данных"
                oo_login = self.get_id_oo(data_sheet["A" + str(row)].value, year)
                for id_description, description in zip([1, 2, 3, 4, 5, 6, 7],
                                                       [description_1, description_2, description_3, description_4,
                                                        description_5, description_6, description_7]):
                    self._cur.execute(f"INSERT INTO oo_description_of_career_guidance ("
                                      f" id_oo,"
                                      f" id_description_of_career_guidance,"
                                      f" value)"
                                      f" VALUES ({oo_login}, {id_description}, '{description}')")
            print("Таблица oo_description_of_career_guidance заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_levels_of_the_educational_program(self):
        try:
            for values in ["Начальное общее образование",
                           "Основное общее образование",
                           "Среднее общее образование",
                           "Среднее профессиональное образование"]:
                self._cur.execute(
                    f"INSERT INTO levels_of_the_educational_program (educational_program) VALUES ('{values.replace(' ', '_')}')")
            print("Таблица levels_of_the_educational_program заполненна")

        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_oo_levels_of_the_educational_program(self, year):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            for row in range(2, data_sheet.max_row + 1):
                lvl1 = data_sheet["J" + str(row)].value if data_sheet["J" + str(row)].value else "Нет_данных"
                lvl2 = data_sheet["K" + str(row)].value if data_sheet["K" + str(row)].value else "Нет_данных"
                lvl3 = data_sheet["L" + str(row)].value if data_sheet["L" + str(row)].value else "Нет_данных"
                lvl4 = data_sheet["M" + str(row)].value if data_sheet["M" + str(row)].value else "Нет_данных"
                oo_login = self.get_id_oo(data_sheet["A" + str(row)].value, year)
                for id_levels_of_the_educational_program, value in zip([1, 2, 3, 4],
                                                                       [lvl1, lvl2, lvl3, lvl4]):
                    self._cur.execute(f"INSERT INTO oo_levels_of_the_educational_program ("
                                      f" id_oo,"
                                      f" id_levels_of_the_educational_program,"
                                      f" value)"
                                      f" VALUES ({oo_login}, {id_levels_of_the_educational_program}, '{value}')")
            print("Таблица oo_levels_of_the_educational_program заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_percentage_of_parents_attending_parentteacher_meeting(self, year):
        try:
            data = openpyxl.reader.excel.load_workbook(
                filename="../excel/Сбор контекстных данных об ОО и участниках ВПР 2021.xlsx", data_only=True)
            data_sheet = data.active
            for row in range(2, data_sheet.max_row + 1):
                lvl1 = data_sheet["CB" + str(row)].value if data_sheet["CB" + str(row)].value else 0
                lvl2 = data_sheet["CC" + str(row)].value if data_sheet["CC" + str(row)].value else 0
                lvl3 = data_sheet["CF" + str(row)].value if data_sheet["CF" + str(row)].value else 0
                oo_login = self.get_id_oo(data_sheet["A" + str(row)].value, year)
                for id_levels_of_the_educational_program, value in zip([1, 2, 3],
                                                                       [lvl1, lvl2, lvl3]):
                    self._cur.execute(f"INSERT INTO percentage_of_parents_attending_parentteacher_meeting ("
                                      f" id_oo,"
                                      f" id_levels_of_the_educational_program,"
                                      f" value)"
                                      f" VALUES ({oo_login}, {id_levels_of_the_educational_program}, {value})")
            print("Таблица percentage_of_parents_attending_parentteacher_meeting заполненна")
            data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_parallels(self):
        try:
            for value in [4, 5, 6, 7, 8, 10, 11]:
                self._cur.execute(
                    f"INSERT INTO parallels (parallel) VALUES ({value})")
            print("Таблица parallels заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_oo_parallels(self, year):
        try:
            all_parallels = glob(f"../excel/vpr_results/{year}/*")
            parallels_schools = {}
            for path in all_parallels:
                parallel = path[path.rfind("\\") + 1:]
                if parallel == "Сбор контекстных данных об ОО (2022).xlsx":
                    continue
                parallels_schools[int(parallel)] = set()
                subj_in_parallel = path.replace("\\", "/") + "/*"
                files = glob(subj_in_parallel)
                for file in files:
                    if year == 2020:
                        df = VPR20(file)
                    elif year == 2022:
                        print(file)
                        df = VPR22(file)
                    else:
                        df = VPR(file)
                    parallels_schools[int(parallel)] |= set(df.get_unique_schools())
            for parallel in parallels_schools:
                for login in parallels_schools[parallel]:
                    id_oo = self.get_id_oo(login, year)
                    self._cur.execute(
                        f"INSERT INTO oo_parallels (parallel, id_oo) VALUES ({parallel}, {id_oo})")
            print("Таблица oo_parallels заполненна")

        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_subjects(self):
        try:
            subjects = ["Английский_язык", "Биология", "География", "История", "Математика", "Немецкий_язык",
                        "Обществознание", "Окружающий_мир", "Русский_язык", "Физика", "Французский_язык", "Химия"]
            for subject in subjects:
                self._cur.execute(
                    f"INSERT INTO subjects (subject_name) VALUES ('{subject}')")

            print("Таблица subjects заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_textbooks(self):
        try:
            books_data = openpyxl.reader.excel.load_workbook(
                filename="../excel/books.xlsx", data_only=True)
            books_sheet = books_data.active
            subj_dict = {"bio": "Биология", "de": "Немецкий_язык", "en": "Английский_язык", "fi": "Физика",
                         "fr": "Французский_язык", "geo": "География", "hi": "Химия", "him": "Химия", "is": "История",
                         "ma": "Математика", "ob": "Обществознание", "om": "Окружающий_мир", "ru": "Русский_язык"}
            for row in range(1, books_sheet.max_row + 1):
                book_key = books_sheet["A" + str(row)].value
                book_name = books_sheet["B" + str(row)].value
                for subj_key in subj_dict:
                    if subj_key in book_key:
                        subj_name = subj_dict[subj_key]
                        id_subjects = self.get_id_subjects(subj_name)

                self._cur.execute(
                    f"INSERT INTO textbooks (id_subjects, key, name) VALUES ({id_subjects}, '{book_key}', '{book_name.strip()}')")
            for id_subjects_ in range(1, 13):
                self._cur.execute(
                    f"INSERT INTO textbooks (id_subjects, key, name) VALUES ({id_subjects_}, '{'nan'}', '{'Нет данных'}')")
            print("Таблица textbooks заполненна")
            books_data.close()
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_classes(self, year):
        try:
            all_parallels = glob(f"../excel/vpr_results/{year}/*")
            for path in all_parallels:
                result_dict = {}
                parallel = path[path.rfind("\\") + 1:]
                if parallel == "Сбор контекстных данных об ОО (2022).xlsx":
                    continue
                parallel = int(parallel)
                subj_in_parallel = path.replace("\\", "/") + "/*"
                files = glob(subj_in_parallel)
                for file in files:
                    if year == 2020:
                        df = VPR20(file)
                    elif year == 2022:
                        df = VPR22(file)
                    else:
                        df = VPR(file)
                    dict_schools_and_liters = df.get_dict_schools_liters()
                    for school in dict_schools_and_liters:
                        if result_dict.get(school) is None:
                            result_dict[school] = dict_schools_and_liters[school]
                        else:
                            for liter in dict_schools_and_liters[school]:
                                if liter not in result_dict[school]:
                                    result_dict[school].append(liter)
                for login in result_dict:
                    id_oo = self.get_id_oo(login, year)
                    id_oo_parallels = self.get_id_oo_parallels(parallel, id_oo)
                    for liter in result_dict[login]:
                        try:
                            self._cur.execute(
                                f"INSERT INTO classes (id_oo_parallels, liter) VALUES ({id_oo_parallels}, '{liter}')")
                        except:
                            print(parallel, subj_in_parallel, liter)
            print("Таблица classes заполненна")

        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_classes_textbooks(self):
        try:
            all_parallels = glob("excel/vpr_results/*")
            result_dict = {}
            for path in all_parallels:
                parallel = int(path[path.rindex("/") + 1:])
                subj_in_parallel = path.replace("/", "/") + "/*"
                files = glob(subj_in_parallel)
                for file in files:
                    print(file.replace("/", "/"))
                    df = VPR(file.replace("/", "/"))
                    id_subjects = self.get_id_subjects(df.get_subj_name())
                    dict_books = df.get_dict_schools_liters_books()
                    if result_dict.get(parallel) is None:
                        result_dict[parallel] = {id_subjects: dict_books}
                    else:
                        if result_dict[parallel].get(id_subjects) is None:
                            result_dict[parallel][id_subjects] = dict_books
                        else:
                            for school in dict_books:
                                if result_dict[parallel][id_subjects].get(school) is None:
                                    result_dict[parallel][id_subjects][school] = dict_books[school]
                                else:
                                    for liter in dict_books[school]:
                                        if result_dict[parallel][id_subjects][school].get(liter) is None:
                                            result_dict[parallel][id_subjects][school][liter] = dict_books[school][
                                                liter]
            for parallel in result_dict:
                for id_subjects in result_dict[parallel]:
                    for login_school in result_dict[parallel][id_subjects]:
                        id_oo_parallels = self.get_id_oo_parallels(parallel, self.get_id_oo(login_school))
                        for liter in result_dict[parallel][id_subjects][login_school]:
                            book_name = result_dict[parallel][id_subjects][login_school][liter]
                            id_classes = self.get_id_classes(id_oo_parallels, liter)
                            id_textbooks = self.get_id_textbooks(id_subjects, book_name.strip())
                            self._cur.execute(
                                f"INSERT INTO classes_textbooks (id_classes, id_textbooks, id_oo_parallels) "
                                f"VALUES ({id_classes}, {id_textbooks}, {id_oo_parallels})")
            print("Таблица classes_textbooks заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_students(self, year):
        try:
            all_parallels = glob(f"../excel/vpr_results/{year}/*")
            result_dict = {}  # {parallel: {school: {liter: {students: gender}}}}
            for path in all_parallels:
                parallel = path[path.rfind("\\") + 1:]
                if parallel == "Сбор контекстных данных об ОО (2022).xlsx":
                    continue
                parallel = int(parallel)
                subj_in_parallel = path.replace("\\", "/") + "/*"
                files = glob(subj_in_parallel)
                for file in files:
                    print(file.replace("\\", "/"))
                    if year == 2020:
                        df = VPR20(file.replace("\\", "/"))
                    elif year == 2022:
                        df = VPR22(file.replace("\\", "/"))
                    else:
                        df = VPR(file.replace("\\", "/"))
                    dict_schools_liters_students = df.get_dict_schools_liters_students()  # {school: {liter: {students: gender}}}
                    if result_dict.get(parallel) is None:
                        result_dict[parallel] = dict_schools_liters_students
                    else:
                        for school in dict_schools_liters_students:
                            if result_dict[parallel].get(school) is None:
                                result_dict[parallel][school] = dict_schools_liters_students[school]
                            else:
                                for liter in dict_schools_liters_students[school]:
                                    if result_dict[parallel][school].get(liter) is None:
                                        result_dict[parallel][school][liter] = dict_schools_liters_students[school][
                                            liter]
                                    else:
                                        for student in dict_schools_liters_students[school][liter]:
                                            if result_dict[parallel][school][liter].get(student) is None:
                                                gender = dict_schools_liters_students[school][liter][student]
                                                result_dict[parallel][school][liter][student] = gender
            count = 0
            for parallel in result_dict:
                for school in result_dict[parallel]:
                    id_oo_parallels = self.get_id_oo_parallels(parallel, self.get_id_oo(school, year))
                    for liter in result_dict[parallel][school]:
                        count += len(result_dict[parallel][school][liter])
                        id_classes = self.get_id_classes(id_oo_parallels, liter)
                        for student in result_dict[parallel][school][liter]:
                            gender = result_dict[parallel][school][liter][student]
                            self._cur.execute(
                                f"INSERT INTO students (id_oo_parallels, id_classes, gender, student_number) "
                                f"VALUES ({id_oo_parallels}, {id_classes}, '{gender}', '{student}')")
            print("Таблица students заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_oo_parallels_subjects(self, year):
        try:
            all_parallels = glob(f"excel/vpr_results/{year}/*")
            for path in all_parallels:
                parallel = path[path.rfind("\\") + 1:]
                if parallel == "Сбор контекстных данных об ОО (2022).xlsx":
                    continue
                parallel = int(parallel)
                subj_in_parallel = path.replace("\\", "/") + "/*"
                files = glob(subj_in_parallel)
                for file in files:
                    subj_name = file[file.rfind("\\") + 1:].replace(".xlsx", "")
                    if year == 2020:
                        df = VPR20(file.replace("\\", "/"))
                        mark_three, mark_four, mark_five = df.get_balls(subj_name, parallel)
                        id_subjects = self.get_id_subjects(subj_name)
                    else:
                        df = VPR(file.replace("\\", "/"))
                        mark_three, mark_four, mark_five = df.get_translation_scale()
                        id_subjects = self.get_id_subjects(df.get_subj_name())
                    unic_schools = df.get_unic_schools()
                    for school in unic_schools:
                        id_oo = self.get_id_oo(school, year)
                        id_oo_parallels = self.get_id_oo_parallels(parallel, id_oo)
                        self._cur.execute(f"INSERT INTO oo_parallels_subjects "
                                          f"(id_subjects, id_oo_parallels, mark_three, mark_four, mark_five) "
                                          f"VALUES "
                                          f"({id_subjects}, {id_oo_parallels}, {mark_three},"
                                          f" {mark_four}, {mark_five})")
            print("Таблица oo_parallels_subjects заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def thread_fill_result_for_task(self, year, file, parallel):
        try:

            print(file.replace("/", "/"))
            if year == 2020:
                df = VPR20(file.replace("/", "/"))
            else:
                df = VPR(file.replace("/", "/"))
            list_df = df.df_to_list()
            for row in list_df:
                if year == 2020:
                    school, student_number, liter, mark_for_last_semester = row[1], row[3], row[6], row[4]
                    variant = row[7] if row[7] != "отсутствовал" else row[8]
                    marks = row[10:]
                    subj_name = file[file.rfind("\\") + 1:].replace(".xlsx", "")
                    id_subjects = self.get_id_subjects(subj_name)

                else:
                    school, student_number, liter, mark_for_last_semester, variant, marks = row[0], row[2], row[6], \
                                                                                            row[4], row[7], row[9:]
                    id_subjects = self.get_id_subjects(df.get_subj_name())
                id_oo = self.get_id_oo(school.strip(), year)
                id_oo_parallels = self.get_id_oo_parallels(parallel, id_oo)
                id_classes = self.get_id_classes(id_oo_parallels, liter)
                id_students = self.get_id_students(id_oo_parallels, id_classes, student_number)

                id_oo_parallels_subjects = self.get_id_oo_parallels_subjects(id_subjects, id_oo_parallels)

                r = """
                INSERT INTO result_for_task 
                (task_number, id_oo_parallels_subjects, id_students,
                 id_oo_parallels, id_subjects, variant, mark_for_last_semester, mark, parallel,
                 id_distributio_of_tasks_by_positions_of_codifiers) VALUES 
                """
                for task_number, mark in enumerate(marks):
                    id_distributio_of_tasks_by_positions_of_codifiers = \
                        self.get_id_distributio_of_tasks_by_positions_of_codifiers(id_subjects=id_subjects,
                                                                                   parallel=parallel,
                                                                                   task_number=task_number + 1,
                                                                                   year=year)
                    s = f"""
                            ({task_number + 1}, {id_oo_parallels_subjects}, {id_students},
                            {id_oo_parallels}, {id_subjects}, {int(variant)},
                            {int(mark_for_last_semester)}, {int(mark)}, {parallel},
                            {id_distributio_of_tasks_by_positions_of_codifiers}), """
                    r += s
                self._cur.execute(r.rstrip(", "))

        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_result_for_task(self, year):
        try:
            # self._cur.execute("TRUNCATE TABLE result_for_task RESTART IDENTITY cascade;")
            all_parallels = glob(f"excel/vpr_results/{year}/*")
            thread_list = []
            for path in all_parallels:
                parallel = int(path[path.rindex("\\") + 1:])
                subj_in_parallel = path.replace("\\", "/") + "/*"
                files = glob(subj_in_parallel)

                for file in files:
                    config = Config()
                    psql = FillDb(psycopg2.connect(dbname=config.DB_NAME,
                                                   user=config.USER,
                                                   password=config.PASSWORD,
                                                   host=config.HOST,
                                                   port=config.PORT))
                    thread_list.append(
                        threading.Thread(target=psql.thread_fill_result_for_task, args=(year, file, parallel,)))
            for thread in thread_list:
                thread.start()

            if all(thread.is_alive() is False for thread in thread_list):
                print("Таблица result_for_task заполненна")
        except Exception as e:
            print(e)

    def fill_users_oo_logins(self):
        try:
            self._cur.execute("SELECT id_user, login, name, id_role FROM users")
            users = self._cur.fetchall()
            self._cur.execute("SELECT oo_login FROM oo_logins")
            all_oo_logins = self._cur.fetchall()
            for user in users:
                id_user, user_login, name, id_role = user
                if id_role in {1, 2}:  # admin and ministry
                    for oo_login in all_oo_logins:
                        oo_login, = oo_login
                        self._cur.execute(
                            f"INSERT INTO users_oo_logins (id_user, oo_login) VALUES ({id_user}, '{oo_login}')")
                elif id_role == 3:  # municipality
                    self._cur.execute(f"""SELECT DISTINCT oo_login FROM oo 
                        WHERE id_name_of_the_settlement in 
                            (SELECT id_name_of_the_settlement FROM name_of_the_settlement 
                                WHERE id_district in 
                                    (SELECT id_district FROM district 
                                        WHERE district_name = '{name.replace(' ', '_')}'))""")
                    logins = self._cur.fetchall()
                    for oo_login in logins:
                        oo_login, = oo_login
                        self._cur.execute(
                            f"INSERT INTO users_oo_logins (id_user, oo_login) VALUES ({id_user}, '{oo_login}')")
                elif id_role == 4:  # school
                    self._cur.execute(
                        f"INSERT INTO users_oo_logins (id_user, oo_login) VALUES ({id_user}, '{user_login}')")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_kt(self, year):
        try:
            # self._cur.execute("TRUNCATE TABLE kt RESTART IDENTITY cascade;")
            parallels = glob("excel/Описание_работ/*")

            for p in parallels:
                parallel = int(p.replace("excel/Описание_работ\\", ""))
                subjects = glob(p + "/*")
                for s in subjects:
                    subject = s.replace("excel/Описание_работ\\" + str(parallel) + "\\", "")
                    id_subjects = self.get_subject_id(subject)
                    kt_data = openpyxl.reader.excel.load_workbook(
                        filename=s + "/KT.xlsx", data_only=True)
                    kt_sheet = kt_data.active
                    for row in range(2, kt_sheet.max_row + 1):
                        kt_key = kt_sheet["A" + str(row)].value
                        description = kt_sheet["B" + str(row)].value
                        self._cur.execute(f"INSERT INTO kt (kt_key, id_subjects, parallel, description, year) "
                                          f"VALUES ('{str(kt_key).strip()}', {id_subjects}, {parallel}, '{description}', '{year}')")
                    self._cur.execute(f"INSERT INTO kt (kt_key, id_subjects, parallel, description, year) "
                                      f"VALUES ('NULL', {id_subjects}, {parallel}, '', '{year}')")
            print("Таблица kt заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_ks(self, year):
        try:
            # self._cur.execute("TRUNCATE TABLE ks RESTART IDENTITY cascade;")
            parallels = glob("excel/Описание_работ/*")

            for p in parallels:
                parallel = int(p.replace("excel/Описание_работ\\", ""))
                subjects = glob(p + "/*")
                for s in subjects:
                    subject = s.replace("excel/Описание_работ\\" + str(parallel) + "\\", "")
                    id_subjects = self.get_subject_id(subject)

                    if glob(s + "/KS.xlsx"):
                        ks_data = openpyxl.reader.excel.load_workbook(
                            filename=s + "/KS.xlsx", data_only=True)
                        ks_sheet = ks_data.active
                        for row in range(2, ks_sheet.max_row + 1):
                            ks_key = ks_sheet["A" + str(row)].value
                            description = ks_sheet["B" + str(row)].value
                            self._cur.execute(f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) "
                                              f"VALUES ('{str(ks_key).strip()}', {id_subjects}, {parallel}, '{description}', '{year}')")
                        self._cur.execute(f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) "
                                          f"VALUES ('NULL', {id_subjects}, {parallel}, '',  '{year}')")
            self._cur.execute(
                f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) VALUES ('NULL', 1, 11, '', '{year}')")

            self._cur.execute(
                f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) VALUES ('NULL', 6, 7, '', '{year}')")

            self._cur.execute(
                f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) VALUES ('NULL', 6, 11, '', '{year}')")
            self._cur.execute(
                f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) VALUES ('NULL', 11, 7, '', '{year}')")
            self._cur.execute(
                f"INSERT INTO ks (ks_key, id_subjects, parallel, description, year) VALUES ('NULL', 11, 11, '', '{year}')")
            print("Таблица ks заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_distributio_of_tasks_by_positions_of_codifiers(self, year):
        try:
            # self._cur.execute("TRUNCATE TABLE distributio_of_tasks_by_positions_of_codifiers RESTART IDENTITY cascade;")
            parallels = glob("excel/Описание_работ/*")
            for p in parallels:
                parallel = int(p.replace("excel/Описание_работ\\", ""))
                subjects = glob(p + "/*")
                for s in subjects:
                    subject = s.replace("excel/Описание_работ\\" + str(parallel) + "\\", "")
                    id_subjects = self.get_subject_id(subject)

                    all_data = openpyxl.reader.excel.load_workbook(
                        filename=s + "/ALL.xlsx", data_only=True)
                    all_sheet = all_data.active
                    for row in range(2, all_sheet.max_row + 1):
                        task_number_from_kim = all_sheet["A" + str(row)].value
                        task_number = all_sheet["B" + str(row)].value
                        fgos = all_sheet["C" + str(row)].value
                        if fgos:
                            fgos = fgos.replace("- ", "").replace("  ", " ").replace("\n", " ").replace("_x0002_", "")
                        poop_noo = all_sheet["D" + str(row)].value
                        if poop_noo:
                            poop_noo = poop_noo.replace("- ", "").replace("  ", " ").replace("\n", " ").replace(
                                "_x0002_", "")
                        level = all_sheet["E" + str(row)].value
                        if level:
                            if level.upper() == "Б":
                                level = "Базовый"
                            elif level.upper() == "П":
                                level = "Продвинутый"
                            elif level.upper() == "В":
                                level = "Высокий"
                        max_mark = all_sheet["H" + str(row)].value
                        try:
                            self._cur.execute(f"INSERT INTO distributio_of_tasks_by_positions_of_codifiers "
                                              f"(id_subjects, parallel, task_number, task_number_from_kim, "
                                              f"fgos, poop_noo, level, max_mark, year) "
                                              f"VALUES ({id_subjects}, {parallel}, {task_number}, '{task_number_from_kim}',"
                                              f"'{fgos}', '{poop_noo}', '{level}', {max_mark}, '{year}')")
                        except psycopg2.Error as e:
                            print("Ошибка при заполнении БД " + str(e))
            print("Таблица distributio_of_tasks_by_positions_of_codifiers заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_ks_kt(self, year):
        try:
            # self._cur.execute("TRUNCATE TABLE ks_kt RESTART IDENTITY cascade;")
            parallels = glob("excel/Описание_работ/*")
            for p in parallels:
                parallel = int(p.replace("excel/Описание_работ\\", ""))
                subjects = glob(p + "/*")
                for s in subjects:
                    subject = s.replace("excel/Описание_работ\\" + str(parallel) + "\\", "")
                    id_subjects = self.get_subject_id(subject)

                    all_data = openpyxl.reader.excel.load_workbook(
                        filename=s + "/ALL.xlsx", data_only=True)
                    all_sheet = all_data.active
                    for row in range(2, all_sheet.max_row + 1):
                        task_number = all_sheet["B" + str(row)].value
                        id_ks = all_sheet["F" + str(row)].value
                        id_ks = self.get_id_ks(ks_key=id_ks,
                                               id_subjects=id_subjects,
                                               parallel=parallel,
                                               year=year)
                        id_kt = all_sheet["G" + str(row)].value

                        id_kt = self.get_id_kt(kt_key=id_kt,
                                               id_subjects=id_subjects,
                                               parallel=parallel,
                                               year=year)
                        id_distributio_of_tasks_by_positions_of_codifiers = \
                            self.get_id_distributio_of_tasks_by_positions_of_codifiers(id_subjects=id_subjects,
                                                                                       parallel=parallel,
                                                                                       task_number=task_number,
                                                                                       year=year)
                        try:
                            self._cur.execute(f"INSERT INTO ks_kt (id_distributio_of_tasks_by_positions_of_codifiers, "
                                              f"id_subjects, parallel, id_ks, id_kt, task_number) "
                                              f"VALUES ({id_distributio_of_tasks_by_positions_of_codifiers}, "
                                              f"{id_subjects}, {parallel}, {id_ks}, {id_kt}, {task_number})")
                        except psycopg2.Error as e:
                            print("Ошибка при заполнении БД " + str(e))
                            print(id_distributio_of_tasks_by_positions_of_codifiers, id_subjects,
                                  parallel, id_ks, id_kt, task_number)

            print("Таблица ks_kt заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_schools_in_risk(self):
        try:
            self._cur.execute("TRUNCATE TABLE schools_in_risk RESTART IDENTITY cascade;")
            schools = {
                2021: {4: {"Математика": ["sch220150", "sch220175", "sch220198", "sch223197", "sch223610", "sch223615",
                                          "sch223763", "sch223953", "sch224143", "sch224188", "sch224199", "sch224208",
                                          "sch224234", "sch224235", "sch224246", "sch224259", "sch224263", "sch224268",
                                          "sch224313", "sch224332", "sch224353", "sch224361", "sch224395", "sch226062",
                                          "sch226065"],
                           "Русский язык": ["sch220161", "sch220175", "sch220198", "sch223197", "sch223615",
                                            "sch223646",
                                            "sch223687", "sch223763", "sch223953", "sch224143", "sch224188",
                                            "sch224205",
                                            "sch224208", "sch224234", "sch224238", "sch224246", "sch224259",
                                            "sch224263",
                                            "sch224268", "sch224286", "sch224313", "sch224332", "sch224353",
                                            "sch224361",
                                            "sch224362", "sch224397", "sch226062", "sch226065"]},
                       5: {"Математика": ["sch220128", "sch220163", "sch220175", "sch223197", "sch223610", "sch223953",
                                          "sch224188", "sch224208", "sch224246", "sch224263", "sch224268", "sch224286",
                                          "sch224332", "sch224353", "sch224361", "sch226059", "sch226062", "sch226065"],
                           "Русский язык": ["sch220128", "sch220150", "sch220161", "sch220163", "sch220175",
                                            "sch223197",
                                            "sch223615", "sch224208", "sch224246", "sch224263", "sch224286",
                                            "sch224313",
                                            "sch224332", "sch224353", "sch224361", "sch226059", "sch226062",
                                            "sch226065"]}}}
            for year in schools:
                for parallel in schools[year]:
                    for subject in schools[year][parallel]:
                        for school in schools[year][parallel][subject]:
                            self._cur.execute(f"""
                            INSERT INTO schools_in_risk (year, parallel, id_subjects, id_oo) 
                            VALUES ('{year}', {parallel}, {self.get_id_subjects(subject)}, {self.get_id_oo(school)})""")
            print("Таблица schools_in_risk заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def fill_properties(self):
        try:
            properties = [
                "17.1. Организация осуществляет образовательную деятельность по адаптированным образовательным программам (АОП)?",
                "17.2. ОО является отдельной организацией, осуществляющей образовательную деятельность по АОП (например, специальной (коррекционной) ОО?",
                "17.3. Образование обучающихся по АОП организовано в отдельных классах/группах?",
                "17.4. Образование обучающихся по АОП организовано совместно с другими обучающимися?",
                "18. На каких языках осуществляется образовательная деятельность в организации?",
                "19.1. Ведется ли в ОО углубленное изучение отдельных предметов?",
                "19.2. Есть ли предметы, углубленное изучение которых начинается с классов начальной школы (1–4 классов)?",
                "19.3. Есть ли предметы, углубленное изучение которых начинается с классов основной школы (5–9 классов)?",
                "19.4. Есть ли предметы, углубленное изучение которых начинается с классов средней школы (10–11 классов)?",
                "20. Укажите количество обучающихся в ОО в 2021/2022 учебном году (в сумме по всем классам/курсам)",
                "21. Общее количество обучающихся в 2021/2022 учебном году 2006 года рождения",
                "22. Количество обучающихся в 2021/2022 учебном году 2006 г. р., обучающихся по АОП, с ограниченными возможностями здоровья и/или инвалидов (всего)",
                "23. Общее количество обучающихся в 2021/2022 учебном году 2007 года рождения",
                "24. Количество обучающихся в 2021/2022 учебном году 2007 г. р., обучающихся по АОП, с ограниченными возможностями здоровья и/или инвалидов (всего)",
                "25. Сколько всего в ОО учителей/преподавателей?",
                "26. Сколько в ОО учителей, преподающих в 5–9 классах?",
                "27. Сколько всего в ОО учителей/преподавателей первой и высшей категорий?"]
            for property in properties:
                self._cur.execute(f"INSERT INTO properties (name) VALUES ('{property}');")
            print("Таблица properties заполненна")
        except psycopg2.Error as e:
            print("Ошибка при заполнении БД " + str(e))

    def delete_students(self):
        self._cur.execute(f"""
        select id_oo_parallels from oo_parallels 
        where id_oo in 
        (
            select id_oo from oo 
            where year = '2022'
        );""")
        id_oo_parallels_list =self._cur.fetchall()
        for id_oo_parallels, in id_oo_parallels_list:
            print(id_oo_parallels)
            self._cur.execute(f"""
            delete from students 
            where id_oo_parallels = {id_oo_parallels};
            """)



config = Config()
psql = FillDb(psycopg2.connect(dbname=config.DB_NAME,
                               user=config.USER,
                               password=config.PASSWORD,
                               host=config.HOST,
                               port=config.PORT))
# psql.delete_students()
# psql.fill_organisation_status()
# psql.fill_properties()
# psql.dropAllTables()
# psql.createTables()
# psql.fill_district()
# psql.fill_oo_location_type()
# psql.fill_name_of_the_settlement(2022)
# psql.fill_organizational_and_legal_form()
# psql.fill_oo_logins(2022)
# psql.fill_population_of_the_settlement()
# psql.fill_internet_speed()
# psql.fill_the_involvement_of_students_in_additional_education()
# psql.fill_count_of_parents_attending_events()
# psql.fill_count_of_parents_ready_to_help()
# psql.fill_regular_transport_link()
# psql.fill_frequency_of_regular_transport_link()
# psql.fill_possibility_to_get_to_the_oo_by_public_transport()
# psql.fill_oo(2022)
# psql.fill_duration_of_refresher_courses()
# psql.fill_completed_advanced_training_courses_for_teachers(2020)
# psql.fill_description_of_work_with_teachers_taking_advanced_training_courses()
# psql.fill_description_of_career_guidance()
# psql.fill_oo_description_of_career_guidance(2020)
# psql.fill_levels_of_the_educational_program()
# psql.fill_oo_levels_of_the_educational_program(2020)
# psql.fill_percentage_of_parents_attending_parentteacher_meeting(2020)
# psql.fill_parallels()
# psql.fill_oo_parallels(2022)
# psql.fill_subjects()
# psql.fill_textbooks()
# psql.fill_classes(2022)
# psql.fill_classes_textbooks()
psql.fill_students(2022)
# psql.fill_oo_parallels_subjects(2020)
# psql.create_index_on_oo_parallels_subjects()
# psql.create_roles()
# psql.create_users()
# psql.fill_users_oo_logins()
# psql.fill_kt(2020)
# psql.fill_ks(2020)
# psql.fill_distributio_of_tasks_by_positions_of_codifiers(2020)
# psql.fill_result_for_task(2020)
# psql.create_index_on_result_for_task()
# psql.fill_ks_kt(2020)
# psql.fill_schools_in_risk()
