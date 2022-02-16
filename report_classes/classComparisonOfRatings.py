from report_classes.classBaseReport import BaseReport
from openpyxl import Workbook


class ComparisonOfRatings(BaseReport):
    def __init__(self, request, dbase, user):
        super().__init__(request, dbase, user)

    def get_report(self):
        title = "Гистограмма соотвествия отметок за выполненную работу и отметок по журналу"
        x_axis = ""
        y_axis = "Количество участников в %"
        content = self._report_type["name"]

        if self._district["id"] != "all":
            if self._oo["id"] != "all":
                sub_content = {"all_districts": [f"Все муниципалитеты",
                                                 f"ВПР 2021. {self._parallel['name']} класс",
                                                 f"Предмет: {self._subject['name']}"],
                               "district": [f"{self._district['name']}",
                                            f"ВПР 2021. {self._parallel['name']} класс",
                                            f"Предмет: {self._subject['name']}"],
                               "oo": [f"{self._oo['name']}",
                                      f"ВПР 2021. {self._parallel['name']} класс",
                                      f"Предмет: {self._subject['name']}"],
                               "all": [f"Сравнительный анализ",
                                       f"ВПР 2021. {self._parallel['name']} класс",
                                       f"Предмет: {self._subject['name']}"
                                       ]}

                percents = {"all_districts": {},
                            "district": {},
                            "oo": {}}
                id_oo = self._dbase.get_id_oo(oo_login=self._oo["id"],
                                              year=self._year["name"])
                id_oo_parallels = self._dbase.get_id_oo_parallels(parallel=self._parallel['name'],
                                                                  id_oo=id_oo)
                id_oo_parallels_subjects = self._dbase.get_id_oo_parallels_subjects(id_subjects=self._subject["id"],
                                                                                    id_oo_parallels=id_oo_parallels)
                percents_oo = self._dbase.get_comparison_of_ratings(
                    id_oo_parallels_subjects=id_oo_parallels_subjects,
                    id_oo_parallels=id_oo_parallels)

                percents_district = self._dbase.get_comparison_of_ratings_for_all_schools_in_district(
                    id_district=self._district["id"],
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    year=self._year["name"])

                percents_all = self._dbase.get_comparison_of_ratings_for_all_districts(
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    year=self._year["name"])

                percents["oo"]["name"] = self._oo["name"]
                percents["oo"]["value"] = percents_oo
                percents["district"]["name"] = self._district["name"]
                percents["district"]["value"] = percents_district
                percents["all_districts"]["name"] = "Все муниципалитеты"
                percents["all_districts"]["value"] = percents_all

                return {"plot_settings": {"lables": ['Понизили', 'Подтвердили', 'Повысили'],
                                          "content": content,
                                          "sub_content": sub_content,
                                          "title": title,
                                          "x_axis": x_axis,
                                          "y_axis": y_axis},
                        "table_settings": {"titles": ['Группы участников', 'Кол-во участников', '%'],
                                           "fields": ['groups', 'count_of_all_students', '%'],
                                           'groups': ['Понизили (Отметка < Отметка по журналу)',
                                                      'Подтвердили (Отметка = Отметка по журналу)',
                                                      "Повысили (Отметка > Отметка по журналу)",
                                                      "Всего"]},
                        "district": self._district,
                        "oo": self._oo,
                        "percents": percents,
                        }

            elif self._oo["id"] == "all":
                sub_content = {"all_districts": [f"Все муниципалитеты",
                                                 f"ВПР 2021. {self._parallel['name']} класс",
                                                 f"Предмет: {self._subject['name']}"],
                               "district": [f"{self._district['name']}",
                                            f"ВПР 2021. {self._parallel['name']} класс",
                                            f"Предмет: {self._subject['name']}"]}
                percents = {"all_districts": {},
                            "district": {}}

                percents_district = self._dbase.get_comparison_of_ratings_for_all_schools_in_district(
                    id_district=self._district["id"],
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    year=self._year["name"])

                percents_all = self._dbase.get_comparison_of_ratings_for_all_districts(
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    year=self._year["name"])

                percents["district"]["name"] = self._district["name"]
                percents["district"]["value"] = percents_district
                percents["all_districts"]["name"] = "Все муниципалитеты"
                percents["all_districts"]["value"] = percents_all

                return {"plot_settings": {"lables": ['Понизили', 'Подтвердили', 'Повысили'],
                                          "content": content,
                                          "sub_content": sub_content,
                                          "title": title,
                                          "x_axis": x_axis,
                                          "y_axis": y_axis},
                        "table_settings": {"titles": ['Группы участников', 'Кол-во участников', '%'],
                                           'groups': ['Понизили (Отметка < Отметка по журналу)',
                                                      'Подтвердили (Отметка = Отметка по журналу)',
                                                      "Повысили (Отметка > Отметка по журналу)",
                                                      "Всего"]},
                        "district": self._district,
                        "oo": self._oo,
                        "percents": percents,
                        }

        elif self._district["id"] == "all":
            sub_content = {"all_districts": [f"Все муниципалитеты",
                                             f"ВПР 2021. {self._parallel['name']} класс",
                                             f"Предмет: {self._subject['name']}"]}
            if self._oo["id"] == "all":
                percents = {"all_districts": {}}
                percents_all = self._dbase.get_comparison_of_ratings_for_all_districts(
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    year=self._year["name"])

                percents["all_districts"]["name"] = "Все муниципалитеты"
                percents["all_districts"]["value"] = percents_all

                return {"plot_settings": {"lables": ['Понизили', 'Подтвердили', 'Повысили'],
                                          "content": content,
                                          "sub_content": sub_content,
                                          "title": title,
                                          "x_axis": x_axis,
                                          "y_axis": y_axis},
                        "table_settings": {"titles": ['Группы участников', 'Кол-во участников', '%'],
                                           'groups': ['Понизили (Отметка < Отметка по журналу)',
                                                      'Подтвердили (Отметка = Отметка по журналу)',
                                                      "Повысили (Отметка > Отметка по журналу)",
                                                      "Всего"]},
                        "district": self._district,
                        "oo": self._oo,
                        "percents": percents,
                        }

    def export_report(self):
        key = self._request["table_type"]
        report = self.get_report()
        file_name = 'Comparison Of Ratings.xlsx'
        wb = Workbook()
        ws = wb.active
        for index, title in enumerate(report["table_settings"]["titles"]):
            ws.cell(row=1, column=index + 1, value=title)
        row = 2
        groups = ["Понизили", "Подтвердили", "Повысили", "Всего"]
        for index, group in enumerate(groups):
            col = 1
            ws.cell(row=row, column=col, value=report["table_settings"]["groups"][index])
            for row_key in ["count_of_students", "%"]:
                col += 1
                ws.cell(row=row, column=col, value=report["percents"][key]["value"][group][row_key])
            row += 1

        return wb, file_name
