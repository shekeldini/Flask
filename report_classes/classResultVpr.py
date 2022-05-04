from report_classes.classBaseReport import BaseReport
from openpyxl import Workbook
from pprint import pprint


class ResultVpr(BaseReport):
    def __init__(self, request, dbase, user):
        super().__init__(request, dbase, user)

    def get_report(self):
        content = self._report_type["name"]
        sub_content = [f"ВПР {' - '.join(self._years['name'].split(','))}. {self._parallel['name']} класс",
                       f"Предмет: {self._subject['name']}"]
        result = {
            "table_settings":
                {
                    "titles": ['Группы участников', 'Кол-во участников', '2', '3', '4', '5', "Средняя отметка",
                               "Качество обученности, %", "Успеваемость, %"],
                    "fields": ["count_of_students", '2', '3', '4', '5', "mean_mark", "quality", "performance"],
                    "values": dict(),
                    "district": self._district,
                    "oo": self._oo
                },
            "content": content,
            "sub_content": sub_content
        }
        if self._district["id"] != "all":
            if self._oo["id"] != "all":
                for year in self._years['name'].split(','):
                    percents = {"all_districts": {},
                                "district": {},
                                "oo": {}}
                    id_oo = self._dbase.get_id_oo(oo_login=self._oo["id"],
                                                  year=year)
                    id_oo_parallels = self._dbase.get_id_oo_parallels(parallel=self._parallel['name'],
                                                                      id_oo=id_oo)
                    id_oo_parallels_subjects = self._dbase.get_id_oo_parallels_subjects(id_subjects=self._subject["id"],
                                                                                        id_oo_parallels=id_oo_parallels)
                    percents_oo = self._dbase.get_result_vpr(
                        id_oo_parallels_subjects=id_oo_parallels_subjects,
                        id_oo_parallels=id_oo_parallels)

                    percents_district = self._dbase.get_result_vpr_for_all_school_in_district(
                        id_district=self._district["id"],
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["name"],
                        year=year)

                    percents_all = self._dbase.get_result_vpr_for_all_districts(
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["name"],
                        year=year)

                    percents["oo"]["name"] = self._oo["name"]
                    percents["oo"]["value"] = percents_oo
                    percents["district"]["name"] = self._district["name"]
                    percents["district"]["value"] = percents_district
                    percents["all_districts"]["name"] = "Все муниципалитеты"
                    percents["all_districts"]["value"] = percents_all
                    result["table_settings"]["values"] = percents
                return result

            elif self._oo["id"] == "all":
                oo_logins_list = self._dbase.get_oo(
                            id_district=self._district["id"],
                            parallel=self._parallel["id"],
                            id_subjects=self._subject["id"],
                            years=self._years['name'].split(','))
                for year in self._years['name'].split(','):
                    percents = {"all_districts": {},
                                "district": {"schools": {}}
                                }
                    percents_district = self._dbase.get_result_vpr_for_all_school_in_district(
                        id_district=self._district["id"],
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["id"],
                        year=year)
                    for oo_login in oo_logins_list:
                        school_name = self._dbase.get_oo_name_by_oo_login(oo_login, year)

                        id_oo = self._dbase.get_id_oo(oo_login=oo_login,
                                                      year=year)
                        id_oo_parallels = self._dbase.get_id_oo_parallels(parallel=self._parallel['name'],
                                                                          id_oo=id_oo)
                        id_oo_parallels_subjects = self._dbase.get_id_oo_parallels_subjects(
                            id_subjects=self._subject["id"],
                            id_oo_parallels=id_oo_parallels)

                        school_percents = self._dbase.get_result_vpr(id_oo_parallels_subjects=id_oo_parallels_subjects,
                                                                     id_oo_parallels=id_oo_parallels)

                        percents["district"]["schools"][school_name] = school_percents

                    percents_all = self._dbase.get_result_vpr_for_all_districts(
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["name"],
                        year=year)

                    percents["district"]["name"] = self._district["name"]
                    percents["district"]["value"] = percents_district
                    percents["all_districts"]["name"] = "Все муниципалитеты"
                    percents["all_districts"]["value"] = percents_all

                    result["table_settings"]["values"] = percents
                return result

        elif self._district["id"] == "all":
            if self._oo["id"] == "all":
                district_list = self._dbase.get_district_list(parallel=self._parallel["id"],
                                                              id_subjects=self._subject["id"],
                                                              years=self._years['name'].split(','))
                for year in self._years['name'].split(','):
                    percents = {
                        "all_districts": {"districts": {}}
                    }

                    percents_all = self._dbase.get_result_vpr_for_all_districts(
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["id"],
                        year=year)

                    for id_district, district_name in district_list:
                        percents_district = self._dbase.get_result_vpr_for_all_school_in_district(
                            id_district=id_district,
                            id_subjects=self._subject["id"],
                            parallel=self._parallel["id"],
                            year=year)
                        percents["all_districts"]["districts"][district_name.replace("_", " ")] = percents_district
                    percents["all_districts"]["name"] = "Все муниципалитеты"
                    percents["all_districts"]["value"] = percents_all

                    result["table_settings"]["values"] = percents
                return result

    def export_report(self):
        report = self.get_report()
        file_name = 'Result Vpr.xlsx'
        wb = Workbook()
        ws = wb.active
        titles = report["table_settings"]["titles"]
        for index in range(2):
            ws.cell(row=1, column=index + 1, value=titles[index])
            ws.merge_cells(start_row=1, start_column=index + 1, end_row=2, end_column=index + 1)

        ws.cell(row=1, column=3, value="Распределение отметок в %")
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)
        for index in range(6, len(titles)):
            ws.cell(row=1, column=index + 1, value=titles[index])
            ws.merge_cells(start_row=1, start_column=index + 1, end_row=2, end_column=index + 1)

        for index in range(2, 6):
            ws.cell(row=2, column=index + 1, value=titles[index])

        if len(report["table_settings"]["values"]) == 3:
            keys = ["all_districts", "district", "oo"]
            dictionary = None
        elif len(report["table_settings"]["values"]) == 2:
            keys = ["all_districts", "district"]
            dictionary = report["table_settings"]["values"]["district"]["schools"]
        else:
            keys = ["all_districts"]
            dictionary = report["table_settings"]["values"]["all_districts"]["districts"]
        row = 3
        for key in keys:
            col = 1
            ws.cell(row=row, column=col, value=report["table_settings"]["values"][key]["name"])
            for field in report["table_settings"]["fields"]:
                col += 1
                ws.cell(row=row, column=col, value=report["table_settings"]["values"][key]["value"][field])
            row += 1
        if dictionary:
            for key, values in dictionary.items():
                col = 1
                ws.cell(row=row, column=col, value=key)
                for field in report["table_settings"]["fields"]:
                    col += 1
                    ws.cell(row=row, column=col, value=values[field])
                row += 1
        return wb, file_name
