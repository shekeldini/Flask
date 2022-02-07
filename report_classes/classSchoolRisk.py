from report_classes.classBaseReport import BaseReport
from openpyxl import Workbook
import datetime


class SchoolsInRisk(BaseReport):
    def __init__(self, request, dbase, user):
        super().__init__(request, dbase, user)

    def get_report(self):
        if self._district['id'] != 'all':
            if self._oo['id'] != 'all':
                res = self._dbase.get_schools_in_risk_for_oo(id_oo=self._oo['id'],
                                                             id_subjects=self._subject['id'],
                                                             parallel=self._parallel['id'])
                return {"type": "oo",
                        'plot_settings': {'lables': [2, 3, 4, 5],
                                          'content': res['school_name'],
                                          'sub_content': [f"ВПР 2021. {self._parallel['name']} класс",
                                                          f"Предмет: {self._subject['name']}"],
                                          'title': 'Сравнение отметок с отметками по журналу',
                                          'x_axis': 'Отметки',
                                          'y_axis': 'Кол-во участников'},
                        'table_settings': {
                            'titles': ['Район', 'Название ОО', 'Кол-во участников', '2', '3', '4', '5', '2', '3', '4',
                                       '5'],
                            'fields': ['2', '3', '4', '5'],
                            'content': 'Необъективные результаты ВПР'},
                        'values': res
                        }

            elif self._oo['id'] == 'all':
                res = self._dbase.get_schools_in_risk_for_district(year=self._year['name'],
                                                                   id_district=self._district['id'],
                                                                   parallel=self._parallel['id'],
                                                                   id_subjects=self._subject['id']
                                                                   )
                return {"type": "district",
                        'table_settings': {
                            'titles': ['Район', 'Название ОО'],
                            'content': 'Необъективные результаты ВПР'},
                        'values': res}

        elif self._district['id'] == "all":
            res = self._dbase.get_schools_in_risk_for_all(year=self._year["name"],
                                                          parallel=self._parallel['id'],
                                                          id_subjects=self._subject["id"]
                                                          )
            return {"type": "district",
                    'table_settings': {
                        'titles': ['Район', 'Название ОО'],
                        'content': 'Необъективные результаты ВПР'},
                    'values': res}

    def export_report(self):
        report = self.get_report()
        file_name = 'Shools In Risk.xlsx'
        wb = Workbook()
        ws = wb.active
        if report["type"] in ("district", "all"):

            for index, title in enumerate(report["table_settings"]["titles"]):
                ws.cell(row=1, column=index + 1, value=title)
            row_count = 2
            for key, school_list in report["values"].items():
                for school in school_list:
                    ws.cell(row=row_count, column=1, value=key)
                    ws.cell(row=row_count, column=2, value=school)
                    row_count += 1
        elif report["type"] == "oo":
            titles = report["table_settings"]["titles"]
            for index in range(3):
                ws.cell(row=1, column=index + 1, value=titles[index])
                ws.merge_cells(start_row=1, start_column=index + 1, end_row=2, end_column=index + 1)

            ws.cell(row=1, column=4, value="Отметка по школе")
            ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)

            ws.cell(row=1, column=8, value="Отметка по ВПР")
            ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=11)

            for index in range(3, 11):
                ws.cell(row=2, column=4 + index - 3, value=titles[index])

            ws.cell(row=3, column=1, value=report["values"]["district_name"])
            ws.cell(row=3, column=2, value=report["values"]["school_name"])
            ws.cell(row=3, column=3, value=report["values"]["count_of_students"])
            for index, value in enumerate(list(report["values"]["last_semester_results"].values())):
                ws.cell(row=3, column=4 + index, value=value)

            for index, value in enumerate(report["values"]["vpr_results"].values()):
                ws.cell(row=3, column=8 + index, value=value)

        return wb, file_name
