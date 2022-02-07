from report_classes.classBaseReport import BaseReport
from openpyxl import Workbook


class WorkDescription(BaseReport):
    def __init__(self, request, dbase, user):
        super().__init__(request, dbase, user)

    def get_report(self):
        if self._district["id"] != "all":
            if self._oo["id"] != "all":
                oo = self._dbase.get_task_description_for_oo(id_oo_parallels_subjects=self._subject["id"])

                district = self._dbase.get_task_description_for_district(id_district=self._district["id"],
                                                                         id_subjects=self._dbase.get_subject_id(self._subject["name"]),
                                                                         parallel=self._parallel["name"],
                                                                         year=self._year["name"])

                all_ = self._dbase.get_task_description_for_all(id_subjects=self._dbase.get_subject_id(self._subject["name"]),
                                                                parallel=self._parallel["name"],
                                                                year=self._year["name"])
                return {"table_settings": {"titles": ["Номер задания",
                                                      "Блоки ПООП обучающийся научится "
                                                      "/ получит возможность научиться "
                                                      "или проверяемые требования (умения) "
                                                      "в соответствии с ФГОС (ФК ГОС)", "Макс балл",
                                                      "Все муниципалитеты", self._district["name"], self._oo["name"]],
                                           "title": "Описание контрольных измерительных материалов"},
                        "values_array": {"oo": {"values": oo},
                                         "district": {"values": district},
                                         "all": {"values": all_}}}

            elif self._oo["id"] == "all":
                district = self._dbase.get_task_description_for_district(id_district=self._district["id"],
                                                                         id_subjects=self._subject["id"],
                                                                         parallel=self._parallel["id"],
                                                                         year=self._year["name"])

                all_ = self._dbase.get_task_description_for_all(id_subjects=self._subject["id"],
                                                                parallel=self._parallel["id"],
                                                                year=self._year["name"])
                return {"table_settings": {"titles": ["Номер задания",
                                                      "Блоки ПООП обучающийся научится "
                                                      "/ получит возможность научиться "
                                                      "или проверяемые требования (умения) "
                                                      "в соответствии с ФГОС (ФК ГОС)", "Макс балл",
                                                      "Все муниципалитеты", self._district["name"]],
                                           "title": "Описание контрольных измерительных материалов"},
                        "values_array": {"district": {"values": district},
                                         "all": {"values": all_}}}

        elif self._district["id"] == "all":
            if self._oo["id"] == "all":
                all_ = self._dbase.get_task_description_for_all(id_subjects=self._subject["id"],
                                                                parallel=self._parallel["id"],
                                                                year=self._year["name"])

                return {"table_settings": {"titles": ["Номер задания",
                                                      "Блоки ПООП обучающийся научится "
                                                      "/ получит возможность научиться "
                                                      "или проверяемые требования (умения) "
                                                      "в соответствии с ФГОС (ФК ГОС)", "Макс балл",
                                                      self._district["name"]],
                                           "title": "Описание контрольных измерительных материалов"},
                        "values_array": {"all": {"values": all_}}}

    def export_report(self):
        report = self.get_report()
        file_name = 'Task Description.xlsx'
        wb = Workbook()
        ws = wb.active
        titles = report["table_settings"]["titles"]
        for index in range(3):
            ws.cell(row=1, column=index + 1, value=titles[index])
            ws.merge_cells(start_row=1, start_column=index + 1, end_row=2, end_column=index + 1)
        column = 4
        for index in range(3, len(titles)):

            ws.cell(row=1, column=column, value=titles[index])
            ws.merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + 3)
            for i, category in enumerate(["Выполнили кол-во", "Не выполнили кол-во", "Выполнили в %", "Не выполнили в %"]):
                ws.cell(row=2, column=column + i, value=category)
            column += 4
        for i in range(1, len(report["values_array"]["all"]["values"]) + 1):
            ws.cell(row=2 + i, column=1, value=report["values_array"]["all"]["values"][i]["task_number_from_kim"])
            ws.cell(row=2 + i, column=2, value=report["values_array"]["all"]["values"][i]["text"])
            ws.cell(row=2 + i, column=3, value=report["values_array"]["all"]["values"][i]["max_mark"])
            ws.cell(row=2 + i, column=4, value=report["values_array"]["all"]["values"][i]["values"]["Выполнили"]["count"])
            ws.cell(row=2 + i, column=5, value=report["values_array"]["all"]["values"][i]["values"]["Не выполнили"]["count"])
            ws.cell(row=2 + i, column=6, value=report["values_array"]["all"]["values"][i]["values"]["Выполнили"]["%"])
            ws.cell(row=2 + i, column=7, value=report["values_array"]["all"]["values"][i]["values"]["Не выполнили"]["%"])
            if len(report["values_array"]) > 1:
                ws.cell(row=2 + i, column=8,
                        value=report["values_array"]["district"]["values"][i]["values"]["Выполнили"]["count"])
                ws.cell(row=2 + i, column=9,
                        value=report["values_array"]["district"]["values"][i]["values"]["Не выполнили"]["count"])
                ws.cell(row=2 + i, column=10,
                        value=report["values_array"]["district"]["values"][i]["values"]["Выполнили"]["%"])
                ws.cell(row=2 + i, column=11,
                        value=report["values_array"]["district"]["values"][i]["values"]["Не выполнили"]["%"])

            if len(report["values_array"]) > 2:
                ws.cell(row=2 + i, column=12,
                        value=report["values_array"]["oo"]["values"][i]["values"]["Выполнили"]["count"])
                ws.cell(row=2 + i, column=13,
                        value=report["values_array"]["oo"]["values"][i]["values"]["Не выполнили"]["count"])
                ws.cell(row=2 + i, column=14,
                        value=report["values_array"]["oo"]["values"][i]["values"]["Выполнили"]["%"])
                ws.cell(row=2 + i, column=15,
                        value=report["values_array"]["oo"]["values"][i]["values"]["Не выполнили"]["%"])
        return wb, file_name
