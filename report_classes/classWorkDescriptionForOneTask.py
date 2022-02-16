from report_classes.classBaseReport import BaseReport
from openpyxl import Workbook


class WorkDescriptionForOneTask(BaseReport):
    def __init__(self, request, dbase, user):
        super().__init__(request, dbase, user)
        self.task = request["task"]

    def get_report(self):
        if self._district["id"] != "all":
            if self._oo["id"] != "all":
                id_oo = self._dbase.get_id_oo(oo_login=self._oo["id"],
                                              year=self._year["name"])
                id_oo_parallels = self._dbase.get_id_oo_parallels(parallel=self._parallel['name'],
                                                                  id_oo=id_oo)
                id_oo_parallels_subjects = self._dbase.get_id_oo_parallels_subjects(id_subjects=self._subject["id"],
                                                                                    id_oo_parallels=id_oo_parallels)

                oo = self._dbase.get_task_description_for_one_task_for_oo(id_oo_parallels_subjects=id_oo_parallels_subjects,
                                                                          task_number=self.task["id"])

                district = self._dbase.get_task_description_for_one_task_for_district(
                    id_district=self._district["id"],
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    task_number=self.task["id"],
                    year=self._year["name"])

                all_ = self._dbase.get_task_description_for_one_task_for_all(
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    task_number=self.task["id"],
                    year=self._year["name"])
                return {"table_settings": {"titles": ["Все муниципалитеты", self._district["name"], self._oo["name"]],
                                           "title": "Описание контрольных измерительных материалов"},
                        "values_array": {"oo": {"values": oo},
                                         "district": {"values": district},
                                         "all": {"values": all_}}}

            elif self._oo["id"] == "all":
                district = self._dbase.get_task_description_for_one_task_for_district(
                    id_district=self._district["id"],
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    task_number=self.task["id"],
                    year=self._year["name"])

                all_ = self._dbase.get_task_description_for_one_task_for_all(
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    task_number=self.task["id"],
                    year=self._year["name"])
                return {"table_settings": {"titles": ["Все муниципалитеты", self._district["name"]],
                                           "title": "Описание контрольных измерительных материалов"},
                        "values_array": {"district": {"values": district},
                                         "all": {"values": all_}}}

        elif self._district["id"] == "all":
            if self._oo["id"] == "all":
                all_ = self._dbase.get_task_description_for_one_task_for_all(
                    id_subjects=self._subject["id"],
                    parallel=self._parallel["name"],
                    task_number=self.task["id"],
                    year=self._year["name"])

                return {"table_settings": {"titles": [self._district["name"]],
                                           "title": "Описание контрольных измерительных материалов"},
                        "values_array": {"all": {"values": all_}}}

    def export_report(self):
        report = self.get_report()
        file_name = 'Task Description For One Task.xlsx'
        wb = Workbook()
        ws = wb.active

        if len(report["values_array"]) == 3:
            colspan = 2
            keys = ["all", "district", "oo"]
        elif len(report["values_array"]) == 2:
            colspan = 1
            keys = ["all", "district"]
        else:
            colspan = 0
            keys = ["all"]
        row = 1
        for task, value in report["values_array"]["all"]["values"].items():
            for title, value_for_title in zip(["№", "Умения, виды деятельности", "Уровень сложности", "Максимальный балл"],
                                              [value["task_number_from_kim"], value["text"], value["level"], value["max_mark"]]):
                ws.cell(row=row, column=1, value=title)
                ws.cell(row=row, column=2, value=value_for_title)
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + colspan)
                row += 1

            ws.cell(row=row, column=1, value="Проверяемые элементы содержания")
            ks_row = row
            for index, ks in enumerate(value["ks"]):
                ws.cell(row=row, column=2, value=f"{index + 1}) {ks}")
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + colspan)
                row += 1
            ws.merge_cells(start_row=ks_row, start_column=1, end_row=row - 1, end_column=1)

            ws.cell(row=row, column=1, value="Проверяемые требования")
            kt_row = row
            for index, kt in enumerate(value["kt"]):
                ws.cell(row=row, column=2, value=f"{index + 1}) {kt}")
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + colspan)
                row += 1
            ws.merge_cells(start_row=kt_row, start_column=1, end_row=row - 1, end_column=1)

            for index, title in enumerate(report["table_settings"]["titles"]):
                ws.cell(row=row, column=2 + index, value=title)

            ws.cell(row=row + 1, column=1, value="Выполнили кол-во")
            ws.cell(row=row + 2, column=1, value="Не выполнили кол-во")
            ws.cell(row=row + 3, column=1, value="Выполнили в %")
            ws.cell(row=row + 4, column=1, value="Не выполнили %")
            for index, key in enumerate(keys):
                ws.cell(row=row + 1, column=2 + index,
                        value=report["values_array"][key]["values"][task]["values"]["Выполнили"]["count"])
                ws.cell(row=row + 2, column=2 + index,
                        value=report["values_array"][key]["values"][task]["values"]["Не выполнили"]["count"])
                ws.cell(row=row + 3, column=2 + index,
                        value=report["values_array"][key]["values"][task]["values"]["Выполнили"]["%"])
                ws.cell(row=row + 4, column=2 + index,
                        value=report["values_array"][key]["values"][task]["values"]["Не выполнили"]["%"])

        return wb, file_name
