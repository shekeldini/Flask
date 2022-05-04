from report_classes.classBaseReport import BaseReport
from openpyxl import Workbook


class StatisticsOfMarks(BaseReport):
    def __init__(self, request, dbase, user):
        super().__init__(request, dbase, user)

    def get_report(self):
        title = "Общая гистограмма отметок"
        x_axis = "Отметка"
        y_axis = "% участников"
        content = self._report_type["name"]
        sub_content = [f"ВПР {' - '.join(self._years['name'].split(','))}. {self._parallel['name']} класс",
                       f"Предмет: {self._subject['name']}"]
        result = {
            "plot_settings": {
                "lables": [2, 3, 4, 5],
                "content": content,
                "sub_content": sub_content,
                "title": title,
                "x_axis": x_axis,
                "y_axis": y_axis,
                "type": None,

            },
            "table_settings": {
                "titles": ['Группы участников', 'Кол-во участников', '2', '3', '4', '5'],
                "fields": ['count_of_all_students', '2', '3', '4', '5'],
                "values": dict()
            },
            "district": self._district,
            "oo": self._oo,
            "percents": dict(),
        }
        if self._district["id"] != "all":
            if self._oo["id"] != "all":
                for year in self._years['name'].split(','):
                    percents = {
                        "all_districts": {},
                        "district": {},
                        "oo": {}
                    }
                    count_of_all_students = {
                        "all_districts": {},
                        "district": {},
                        "oo": {}
                    }
                    id_oo = self._dbase.get_id_oo(oo_login=self._oo["id"],
                                                  year=year)
                    id_oo_parallels = self._dbase.get_id_oo_parallels(parallel=self._parallel['name'],
                                                                      id_oo=id_oo)
                    id_oo_parallels_subjects = self._dbase.get_id_oo_parallels_subjects(id_subjects=self._subject["id"],
                                                                                        id_oo_parallels=id_oo_parallels)
                    percents_oo, count_of_all_students_oo = self._dbase.get_count_students_mark(
                        id_oo_parallels_subjects=id_oo_parallels_subjects,
                        id_oo_parallels=id_oo_parallels)

                    percents_district, count_of_all_students_district = self._dbase.get_count_students_mark_for_all_school_in_district(
                        id_district=self._district["id"],
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["name"],
                        year=year)

                    percents_all, count_of_all_students_all = self._dbase.get_count_students_mark_for_all_districts(
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["name"],
                        year=year)

                    percents["oo"]["name"] = self._oo["name"]
                    percents["oo"]["value"] = percents_oo
                    percents["district"]["name"] = self._district["name"]
                    percents["district"]["value"] = percents_district
                    percents["all_districts"]["name"] = "Все муниципалитеты"
                    percents["all_districts"]["value"] = percents_all

                    count_of_all_students["oo"]["name"] = self._oo["name"]
                    count_of_all_students["oo"]["value"] = count_of_all_students_oo
                    count_of_all_students["district"]["name"] = self._district["name"]
                    count_of_all_students["district"]["value"] = count_of_all_students_district
                    count_of_all_students["all_districts"]["name"] = "Все муниципалитеты"
                    count_of_all_students["all_districts"]["value"] = count_of_all_students_all

                    result["table_settings"]["values"][year] = {
                        'groups': [percents[x]["name"] for x in percents],
                        "count_of_all_students": [count_of_all_students[x]["value"] for x in count_of_all_students],
                        "2": [percents[x]["value"][2] for x in percents],
                        "3": [percents[x]["value"][3] for x in percents],
                        "4": [percents[x]["value"][4] for x in percents],
                        "5": [percents[x]["value"][5] for x in percents]
                    }
                    result["percents"][year] = percents
                result["plot_settings"]["type"] = "oo"

                return result

            elif self._oo["id"] == "all":
                for year in self._years['name'].split(','):
                    percents = {
                        "all_districts": {},
                        "district": {}
                    }
                    count_of_all_students = {
                        "all_districts": {},
                        "district": {}
                    }
                    percents_district, count_of_all_students_district = self._dbase.get_count_students_mark_for_all_school_in_district(
                        id_district=self._district["id"],
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["id"],
                        year=year)

                    percents_all, count_of_all_students_all = self._dbase.get_count_students_mark_for_all_districts(
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["name"],
                        year=year)

                    percents["district"]["name"] = self._district["name"]
                    percents["district"]["value"] = percents_district
                    percents["all_districts"]["name"] = "Все муниципалитеты"
                    percents["all_districts"]["value"] = percents_all

                    count_of_all_students["district"]["name"] = self._district["name"]
                    count_of_all_students["district"]["value"] = count_of_all_students_district
                    count_of_all_students["all_districts"]["name"] = "Все муниципалитеты"
                    count_of_all_students["all_districts"]["value"] = count_of_all_students_all

                    result["table_settings"]["values"][year] = {
                        'groups': [percents[x]["name"] for x in percents],
                        "count_of_all_students": [count_of_all_students[x]["value"] for x in count_of_all_students],
                        "2": [percents[x]["value"][2] for x in percents],
                        "3": [percents[x]["value"][3] for x in percents],
                        "4": [percents[x]["value"][4] for x in percents],
                        "5": [percents[x]["value"][5] for x in percents]
                    }
                    result["percents"][year] = percents
                result["plot_settings"]["type"] = "district"
                return result

        elif self._district["id"] == "all":
            if self._oo["id"] == "all":
                for year in self._years['name'].split(','):
                    percents = {
                        "all_districts": {}
                    }
                    count_of_all_students = {
                        "all_districts": {}
                    }

                    percents_all, count_of_all_students_all = self._dbase.get_count_students_mark_for_all_districts(
                        id_subjects=self._subject["id"],
                        parallel=self._parallel["id"],
                        year=year)

                    percents["all_districts"]["name"] = "Все муниципалитеты"
                    percents["all_districts"]["value"] = percents_all

                    count_of_all_students["all_districts"]["name"] = "Все муниципалитеты"
                    count_of_all_students["all_districts"]["value"] = count_of_all_students_all
                    result["table_settings"]["values"][year] = {
                        'groups': [percents[x]["name"] for x in percents],
                        "count_of_all_students": [count_of_all_students[x]["value"] for x in count_of_all_students],
                        "2": [percents[x]["value"][2] for x in percents],
                        "3": [percents[x]["value"][3] for x in percents],
                        "4": [percents[x]["value"][4] for x in percents],
                        "5": [percents[x]["value"][5] for x in percents]
                    }
                    result["percents"][year] = percents
                result["plot_settings"]["type"] = "all_districts"
                return result

    def export_report(self):
        report = self.get_report()
        file_name = 'Statistics Of Marks.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=report["table_settings"]["titles"][0])
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        for i, year in enumerate(report["table_settings"]["values"].keys()):
            ws.cell(row=1, column=i * 5 + 2, value=year)
            ws.merge_cells(start_row=1, start_column=i * 5 + 2, end_row=1, end_column=(i + 1) * 5 + 1)

            for index, title in enumerate(report["table_settings"]["titles"][1:], start=1):
                ws.cell(row=2, column=i * 5 + 1 + index, value=title)

            row = 3
            for index in range(len(report["table_settings"]["values"][year]["groups"])):
                ws.cell(row=row, column=1, value=report["table_settings"]["values"][year]["groups"][index])
                col = 1
                for field in report["table_settings"]["fields"]:
                    ws.cell(row=row, column=i * 5 + 1 + col,
                            value=report["table_settings"]["values"][year][field][index])
                    col += 1
                row += 1

        return wb, file_name
