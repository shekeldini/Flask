import psycopg2
from openpyxl import Workbook
from data_base.postgresql import Postgresql
from config import *


class P24(Postgresql):
    def __init__(self, connection):
        super().__init__(connection)

    def get_vpr_result_by_district_parallel_subject(self, id_district, parallel, id_subjects):
        try:
            self._cur.execute(f"""
            select sum from 
                (select id_students, sum(mark) from result_for_task 
                    where id_oo_parallels_subjects in 
                    (
                        select id_oo_parallels_subjects from oo_parallels_subjects 
                        where id_subjects = {id_subjects} 
                        and id_oo_parallels in 
                        (
                            select id_oo_parallels from oo_parallels 
                            where parallel = {parallel}
                            and id_oo in 
                            (
                                select id_oo from oo 
                                where id_name_of_the_settlement in 
                                (
                                    select id_name_of_the_settlement from name_of_the_settlement 
                                    where id_district = {id_district}
                                )
                            )
                        )
                    )
                group by id_students order by sum) as t1; """)
            res = self._cur.fetchall()
            if res:
                return [x[0] for x in res]
            return []

        except psycopg2.Error as e:
            print("Ошибка получения данных из ДБ " + str(e))

    def p24(self):
        d = {"Русский язык": [4, 5, 6, 7, 8],
             "Математика": [4, 5, 6, 7, 8]}
        wb = Workbook()
        ws = wb.active
        row = 1
        ws.cell(row=row, column=1, value="Наименование МОУО")
        ws.cell(row=row, column=2, value="Дисциплина")
        ws.cell(row=row, column=3, value="Оценка функциональности, %")
        row += 1
        for id_district, district_name in self.get_districts(id_user=1, year="2021"):
            district_name = district_name.replace("_", " ")
            for subject_name in d:
                values = []
                for parallel in d[subject_name]:
                    res = self.get_vpr_result_by_district_parallel_subject(id_district=id_district,
                                                                           parallel=parallel,
                                                                           id_subjects=self.get_subject_id(
                                                                               subject_name))
                    low = res[25 * (len(res) + 1) // 100]
                    high = res[75 * (len(res) + 1) // 100]
                    values.append(low / high * 100)
                print(f"{district_name=}, {subject_name=}, {sum(values)/len(values)}")
                ws.cell(row=row, column=1, value=district_name)
                ws.cell(row=row, column=2, value=subject_name)
                ws.cell(row=row, column=3, value=round(sum(values)/len(values), 2))
                row += 1
        wb.save("П25.xlsx")

psql = P24(psycopg2.connect(dbname=DB_NAME, user=USER, password=PASSWORD, host=HOST, port=PORT))
psql.p24()
