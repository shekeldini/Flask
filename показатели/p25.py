from openpyxl import Workbook
from data_base.postgresql import Postgresql


class P25(Postgresql):
    def __init__(self, connection):
        super().__init__(connection)

    def p25_db(self, task_number, id_subjects, parallel, id_district):
        self._cur.execute(f"""
        select task_number, count, sum_all_marks from 
            (select task_number, count(id_result_for_task) as count from result_for_task 
            where task_number = {task_number} 
                and id_subjects = {id_subjects} 
                and id_oo_parallels_subjects in 
                    (
                        select id_oo_parallels_subjects from oo_parallels_subjects where id_subjects = {id_subjects} and id_oo_parallels in 
                        (
                            select id_oo_parallels from oo_parallels where parallel = {parallel} and id_oo in 
                            (
                                select id_oo from oo where id_name_of_the_settlement in 
                                (
                                    select id_name_of_the_settlement from name_of_the_settlement where id_district = {id_district}
                                )
                            )
                        )
                    ) group by task_number) as t1
        left join (
            select task_number, sum(mark) as sum_all_marks from result_for_task 
            where task_number = {task_number} 
                and id_subjects = {id_subjects} 
                and id_oo_parallels_subjects in 
                    (
                        select id_oo_parallels_subjects from oo_parallels_subjects where id_subjects = {id_subjects} and id_oo_parallels in 
                        (
                            select id_oo_parallels from oo_parallels where parallel = {parallel} and id_oo in 
                            (
                                select id_oo from oo where id_name_of_the_settlement in 
                                (
                                    select id_name_of_the_settlement from name_of_the_settlement where id_district = {id_district}
                                )
                            )
                        )
                    ) group by task_number
                ) as t3 using (task_number) group by task_number, count, sum_all_marks
        ;""")
        res = self._cur.fetchall()
        if res:
            return res[0]
        return

    def p25(self):
        d = {"Русский язык": {4: ["15.1", "15.2"],
                              5: ["8", "9", "11"],
                              6: ["9", "10", "11", "12.1", "12.2", "14.1", "14.2"],
                              7: ["11.1", "11.2", "13.1", "13.2"],
                              8: ["6", "7", "8", "10"]},
             "Математика": {4: ["3", "4", "9.1", "9.2", "10"],
                            5: ["8", "10", "11.2", "12.1", "12.2"],
                            6: ["5", "6", "11"],
                            7: ["3", "4", "5", "7", "10", "16"],
                            8: ["3", "6", "10", "11", "15", "16.1", "16.2", "18"]}}
        wb = Workbook()
        ws = wb.active
        row = 1
        ws.cell(row=row, column=1, value="Наименование МОУО")
        ws.cell(row=row, column=2, value="Дисциплина")
        ws.cell(row=row, column=3, value="Оценка функциональности, %")
        row += 1
        for id_district, district_name in self.get_districts(id_user=1, year="2021"):
            district_name = district_name.replace("_", " ")
            for subject_name in d:
                kriteriy = 0
                c = 0
                for parallel in d[subject_name]:
                    values = []
                    for task_number_from_kim in d[subject_name][parallel]:
                        real_ball = 0
                        max_ball = 0
                        id_subjects = self.get_subject_id(subject_name=subject_name)
                        max_mark = self.get_max_mark(subject_name, parallel, task_number_from_kim)
                        if max_mark:
                            task_number = self.get_task_number(task_number_from_kim=task_number_from_kim,
                                                               id_subjects=id_subjects,
                                                               parallel=parallel)
                            if task_number:

                                task_number, count, sum_all_marks = self.p25_db(task_number=task_number,
                                                                              id_subjects=id_subjects,
                                                                              id_district=id_district,
                                                                              parallel=parallel)

                                real_ball += sum_all_marks
                                max_ball += count * max_mark

                            else:
                                print(f"{task_number=} {task_number_from_kim=} {id_subjects=} {parallel=}")
                        else:
                            print(f"{max_mark=} {task_number_from_kim=} {id_subjects=} {parallel=}")
                        values.append(real_ball / max_ball)
                    kriteriy += sum(values) / len(values)
                    c += 1

                ws.cell(row=row, column=1, value=district_name)
                ws.cell(row=row, column=2, value=subject_name)
                ws.cell(row=row, column=3, value=round((kriteriy / c) * 100, 2))
                row += 1

        wb.save("Оценка функциональности.xlsx")
